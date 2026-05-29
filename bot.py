"""
╔═══════════════════════════════════════════════════╗
║       DU - Medicine 2024 — Ultimate Edition v3  ║
║                                                   ║
║  Navigation: Semester → Module → Subject → Lec   ║
╚═══════════════════════════════════════════════════╝
"""

import os, logging, sqlite3, shutil, asyncio, io
from datetime import datetime, time as dtime
from collections import defaultdict
from dotenv import load_dotenv

from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup,
    ReplyKeyboardMarkup, ReplyKeyboardRemove
)
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    CallbackQueryHandler, ContextTypes, filters
)

# ─────────────────────────────────────────────────────
# ⚙️  CONFIG
# ─────────────────────────────────────────────────────
load_dotenv()
TOKEN = os.getenv("BOT_TOKEN")
if not TOKEN:
    raise RuntimeError("❌ BOT_TOKEN مش موجود — حط التوكن في ملف .env")
_admin_env = os.getenv("ADMIN_IDS", "5860401902")
_admin_env = _admin_env.strip()
if not _admin_env or not _admin_env.replace(",","").replace(" ","").isdigit():
    _admin_env = "5860401902"
ADMIN_IDS = set(map(int, _admin_env.split(",")))
DB_FILE   = os.getenv("DB_FILE", "eshra7tab.db")
PAGE_SIZE = 5

# ─────────────────────────────────────────────────────
# 🗺️  NAV CACHE  — maps short int IDs to nav state
#     keeps callback_data well under Telegram's 64-byte limit
# ─────────────────────────────────────────────────────
_nav_store: dict = {}   # id -> (sem, module, ctype, subject)
_nav_next:  int  = 0
_NAV_MAX_SIZE = 2000    # امسح القديم لو الـ cache كبر أوي

def md_escape(text: str) -> str:
    """Escape Telegram Markdown v1 special characters in dynamic/user-supplied text."""
    if not text:
        return text
    for ch in ('*', '_', '`', '['):
        text = text.replace(ch, f'\\{ch}')
    return text

def nav_save(sem, module, ctype, subject="") -> int:
    global _nav_next
    key = (sem, module, ctype, subject)
    for k, v in _nav_store.items():
        if v == key:
            return k
    # لو الـ cache وصل الحد الأقصى، امسح أقدم 500 entry
    if len(_nav_store) >= _NAV_MAX_SIZE:
        oldest_keys = list(_nav_store.keys())[:500]
        for k in oldest_keys:
            del _nav_store[k]
    nid = _nav_next
    _nav_next += 1
    _nav_store[nid] = key
    return nid

def nav_load(nid: int):
    return _nav_store.get(nid, ("", "", "L", ""))

# ─────────────────────────────────────────────────────
# 📚  CURRICULUM  —  Semester → Module → [Subjects]
# ─────────────────────────────────────────────────────
DATA = {
    "Sem 1": {
        "Basic of Health and Life 1": [
            "To Be Added",
        ],
        "Basic of Health and Life 2": [
            "Communication and Medical Ethics",
            "Information Technology",
            "English 1",
        ],
    },
    "Sem 2": {
        "Basic of Disease and Therapy":       ["To Be Added"],
        "Blood and Immune System":            ["To Be Added"],
        "Locomotor 1":                        ["To Be Added"],
        "Clinical Skill and Professionalism": ["To Be Added"],
        "Interprofessional Communication":    ["To Be Added"],
        "English 2":                          ["To Be Added"],
    },
    "Sem 3": {
        "Locomotor 2":                      ["To Be Added"],
        "Circulatory and Lymphatic System": ["To Be Added"],
        "Respiratory System":               ["To Be Added"],
        "Urinary System":                   ["To Be Added"],
        "Medical Report Writing":           ["To Be Added"],
    },
    "Sem 4": {
        "Endocrine and Reproductive System": ["To Be Added"],
        "Digestive and Hepatobiliary":       ["To Be Added"],
        "Metabolism and Nutrition":          ["To Be Added"],
        "Investigation in Medicine":         ["To Be Added"],
    },
}

# ─────────────────────────────────────────────────────
# 🗄️  DATABASE
# ─────────────────────────────────────────────────────
import threading
_db_local = threading.local()

# ─────────────────────────────────────────────────────
# 🗄️  DATABASE
# ─────────────────────────────────────────────────────
def get_db():
    if not getattr(_db_local, "conn", None):
        conn = sqlite3.connect(DB_FILE, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")   # أحسن أداء مع reads/writes متزامنين
        conn.execute("PRAGMA synchronous=NORMAL") # أسرع من FULL مع أمان كافي
        _db_local.conn = conn
    return _db_local.conn

def init_db():
    with get_db() as db:
        db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            user_id   INTEGER PRIMARY KEY,
            name      TEXT,
            username  TEXT,
            joined_at TEXT DEFAULT (datetime('now')),
            last_seen TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS lectures (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            semester     TEXT NOT NULL,
            module       TEXT NOT NULL,
            subject      TEXT NOT NULL,
            title        TEXT NOT NULL,
            content      TEXT NOT NULL,
            file_id      TEXT,
            file_type    TEXT,
            content_type TEXT DEFAULT 'lectures',
            created_at   TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS quizzes (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            lecture_id INTEGER NOT NULL,
            question   TEXT NOT NULL,
            opt_a TEXT, opt_b TEXT, opt_c TEXT, opt_d TEXT,
            answer     TEXT NOT NULL,
            FOREIGN KEY (lecture_id) REFERENCES lectures(id)
        );
        CREATE TABLE IF NOT EXISTS quiz_attempts (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id    INTEGER NOT NULL,
            lecture_id INTEGER NOT NULL,
            score      INTEGER NOT NULL,
            total      INTEGER NOT NULL,
            taken_at   TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS bookmarks (
            user_id    INTEGER,
            lecture_id INTEGER,
            PRIMARY KEY (user_id, lecture_id)
        );
        CREATE TABLE IF NOT EXISTS analytics (
            lecture_id INTEGER PRIMARY KEY,
            views      INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS scheduled_broadcasts (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            message  TEXT NOT NULL,
            send_at  TEXT NOT NULL,
            sent     INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS custom_modules (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            semester  TEXT NOT NULL,
            module    TEXT NOT NULL,
            UNIQUE(semester, module)
        );
        CREATE TABLE IF NOT EXISTS custom_subjects (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            semester  TEXT NOT NULL,
            module    TEXT NOT NULL,
            subject   TEXT NOT NULL,
            UNIQUE(semester, module, subject)
        );
        CREATE TABLE IF NOT EXISTS admins (
            user_id   INTEGER PRIMARY KEY,
            name      TEXT,
            added_at  TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS deleted_hardcoded_subjects (
            semester TEXT NOT NULL,
            module   TEXT NOT NULL,
            subject  TEXT NOT NULL,
            PRIMARY KEY (semester, module, subject)
        );
        CREATE TABLE IF NOT EXISTS exam_banks (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            semester  TEXT NOT NULL,
            exam_type TEXT NOT NULL,
            title     TEXT NOT NULL,
            file_id   TEXT NOT NULL,
            file_type TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now'))
        );
        """)
        # Migrations
        cols = [r[1] for r in db.execute("PRAGMA table_info(lectures)").fetchall()]
        if "content_type" not in cols:
            db.execute("ALTER TABLE lectures ADD COLUMN content_type TEXT DEFAULT 'lectures'")
        ucols = [r[1] for r in db.execute("PRAGMA table_info(users)").fetchall()]
        if "last_seen" not in ucols:
            db.execute("ALTER TABLE users ADD COLUMN last_seen TEXT")
            db.execute("UPDATE users SET last_seen = datetime('now') WHERE last_seen IS NULL")

# ── DB helpers ────────────────────────────────────────
def db_register(user):
    db = get_db()
    cur = db.execute(
        "INSERT OR IGNORE INTO users (user_id,name,username) VALUES (?,?,?)",
        (user.id, user.full_name, user.username or "")
    )
    is_new = cur.rowcount > 0
    db.execute(
        "UPDATE users SET last_seen=datetime('now'), name=?, username=? WHERE user_id=?",
        (user.full_name, user.username or "", user.id)
    )
    db.commit()
    return is_new

def db_save_quiz_attempt(uid, lid, score, total):
    with get_db() as db:
        db.execute(
            "INSERT INTO quiz_attempts (user_id,lecture_id,score,total) VALUES (?,?,?,?)",
            (uid, lid, score, total)
        )

def db_get_quiz_stats(lid):
    """Returns (attempts, avg_score_pct) for a lecture's quiz."""
    with get_db() as db:
        r = db.execute(
            "SELECT COUNT(*) attempts, "
            "ROUND(AVG(CAST(score AS FLOAT)/total)*100, 1) avg_pct "
            "FROM quiz_attempts WHERE lecture_id=?", (lid,)
        ).fetchone()
    return r["attempts"], r["avg_pct"] or 0.0

def db_get_user_quiz_history(uid):
    """Returns all quiz attempts for a user with lecture info."""
    with get_db() as db:
        return db.execute(
            "SELECT qa.score, qa.total, qa.taken_at, l.title, l.subject, l.module "
            "FROM quiz_attempts qa JOIN lectures l ON l.id=qa.lecture_id "
            "WHERE qa.user_id=? ORDER BY qa.taken_at DESC LIMIT 20",
            (uid,)
        ).fetchall()

def db_all_uids():
    with get_db() as db:
        return [r["user_id"] for r in db.execute("SELECT user_id FROM users").fetchall()]

def db_get_user_profile(uid):
    """Returns full profile stats for a user."""
    db = get_db()
    user = db.execute(
        "SELECT * FROM users WHERE user_id=?", (uid,)
    ).fetchone()
    if not user:
        return None
    quiz_attempts = db.execute(
        "SELECT COUNT(*) cnt, "
        "ROUND(AVG(CAST(score AS FLOAT)/total*100),1) avg_pct "
        "FROM quiz_attempts WHERE user_id=? AND total > 0", (uid,)
    ).fetchone()
    bookmarks_count = db.execute(
        "SELECT COUNT(*) FROM bookmarks WHERE user_id=?", (uid,)
    ).fetchone()[0]
    best_quiz = db.execute(
        """SELECT qa.score, qa.total, l.title
           FROM quiz_attempts qa JOIN lectures l ON l.id=qa.lecture_id
           WHERE qa.user_id=? AND qa.total>0
           ORDER BY CAST(qa.score AS FLOAT)/qa.total DESC LIMIT 1""", (uid,)
    ).fetchone()
    return {
        "user":            user,
        "quiz_attempts":   quiz_attempts["cnt"],
        "avg_pct":         quiz_attempts["avg_pct"] or 0,
        "bookmarks_count": bookmarks_count,
        "best_quiz":       best_quiz,
    }

def db_find_user(uid):
    """Find user by Telegram ID — for admin search."""
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE user_id=?", (uid,)).fetchone()
    if not user:
        return None
    quiz_attempts = db.execute(
        "SELECT COUNT(*) cnt, "
        "ROUND(AVG(CAST(score AS FLOAT)/total*100),1) avg_pct "
        "FROM quiz_attempts WHERE user_id=? AND total>0", (uid,)
    ).fetchone()
    bookmarks_count = db.execute(
        "SELECT COUNT(*) FROM bookmarks WHERE user_id=?", (uid,)
    ).fetchone()[0]
    recent_quizzes = db.execute(
        """SELECT qa.score, qa.total, qa.taken_at, l.title
           FROM quiz_attempts qa JOIN lectures l ON l.id=qa.lecture_id
           WHERE qa.user_id=? ORDER BY qa.taken_at DESC LIMIT 5""", (uid,)
    ).fetchall()
    return {
        "user":           user,
        "quiz_attempts":  quiz_attempts["cnt"],
        "avg_pct":        quiz_attempts["avg_pct"] or 0,
        "bookmarks":      bookmarks_count,
        "recent_quizzes": recent_quizzes,
    }

def db_add_lecture(sem, module, subject, title, content, file_id=None, file_type=None, content_type="lectures"):
    with get_db() as db:
        cur = db.execute(
            "INSERT INTO lectures (semester,module,subject,title,content,file_id,file_type,content_type) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (sem, module, subject, title, content, file_id, file_type, content_type)
        )
        return cur.lastrowid

def db_get_lecture(lid):
    with get_db() as db:
        return db.execute("SELECT * FROM lectures WHERE id=?", (lid,)).fetchone()

def _natural_key(row):
    import re
    nums = re.findall(r'\d+', row["title"])
    return (int(nums[0]) if nums else float('inf'), row["title"].lower())

def db_get_lectures(sem, module, subject, page=0, ctype="lectures"):
    db = get_db()
    all_rows = db.execute(
        "SELECT * FROM lectures WHERE semester=? AND module=? AND subject=? AND content_type=?",
        (sem, module, subject, ctype)
    ).fetchall()
    total = len(all_rows)
    all_rows = sorted(all_rows, key=_natural_key)
    rows = all_rows[page * PAGE_SIZE : (page + 1) * PAGE_SIZE]
    return rows, total

def db_subject_count(sem, module, subject):
    with get_db() as db:
        return db.execute(
            "SELECT COUNT(*) FROM lectures WHERE semester=? AND module=? AND subject=?",
            (sem, module, subject)
        ).fetchone()[0]

def db_subject_count_typed(sem, module, subject, ctype="lectures"):
    with get_db() as db:
        return db.execute(
            "SELECT COUNT(*) FROM lectures WHERE semester=? AND module=? AND subject=? AND content_type=?",
            (sem, module, subject, ctype)
        ).fetchone()[0]

def db_module_counts_batch(sem):
    """Return dict {module: count} in one query instead of N queries."""
    db = get_db()
    rows = db.execute(
        "SELECT module, COUNT(*) as cnt FROM lectures WHERE semester=? GROUP BY module",
        (sem,)
    ).fetchall()
    return {r["module"]: r["cnt"] for r in rows}

def db_module_count(sem, module):
    return db_module_counts_batch(sem).get(module, 0)

def db_subject_counts_batch(sem, module, ctype_str):
    """Return dict {subject: count} in one query instead of N queries."""
    db = get_db()
    rows = db.execute(
        "SELECT subject, COUNT(*) as cnt FROM lectures "
        "WHERE semester=? AND module=? AND content_type=? GROUP BY subject",
        (sem, module, ctype_str)
    ).fetchall()
    return {r["subject"]: r["cnt"] for r in rows}

def db_update_lecture(lid, field, value):
    if field not in {"title","content","file_id","file_type","subject","module","semester"}:
        return
    with get_db() as db:
        db.execute(f"UPDATE lectures SET {field}=? WHERE id=?", (value, lid))

def db_delete_lecture(lid):
    with get_db() as db:
        db.execute("DELETE FROM lectures WHERE id=?", (lid,))
        db.execute("DELETE FROM quizzes    WHERE lecture_id=?", (lid,))
        db.execute("DELETE FROM bookmarks  WHERE lecture_id=?", (lid,))
        db.execute("DELETE FROM analytics  WHERE lecture_id=?", (lid,))

def db_sort_lectures_alpha():
    """رتّب كل الملفات أبجدياً داخل كل (semester, module, subject, content_type) بتحديث الـ rowid."""
    db = get_db()
    # جيب كل المجموعات الفريدة
    groups = db.execute(
        "SELECT DISTINCT semester, module, subject, content_type FROM lectures"
    ).fetchall()
    total_moved = 0
    for g in groups:
        sem, mod, sub, ctype = g["semester"], g["module"], g["subject"], g["content_type"]
        lecs = db.execute(
            "SELECT id FROM lectures WHERE semester=? AND module=? AND subject=? AND content_type=? "
            "ORDER BY title COLLATE NOCASE ASC",
            (sem, mod, sub, ctype)
        ).fetchall()
        ids_sorted = [r["id"] for r in lecs]
        if ids_sorted == sorted(ids_sorted):
            continue  # مرتبين أصلاً، مش محتاج نعمل حاجة
        # عمل temp IDs سالبة عشان نتجنب UNIQUE conflicts
        for i, lid in enumerate(ids_sorted):
            db.execute("UPDATE lectures SET id=? WHERE id=?", (-(i+1), lid))
        # رقمهم من أصغر id متاح بالترتيب الأبجدي
        base_id = db.execute("SELECT MAX(id) FROM lectures").fetchone()[0] or 0
        base_id += 1
        new_ids = []
        for i in range(len(ids_sorted)):
            new_id = base_id + i
            db.execute("UPDATE lectures SET id=? WHERE id=?", (new_id, -(i+1)))
            new_ids.append(new_id)
        total_moved += len(new_ids)
    db.commit()
    return total_moved

# ── Exam Banks ────────────────────────────────────────
def db_add_exam_bank(sem, exam_type, title, file_id, file_type):
    db = get_db()
    db.execute(
        "INSERT INTO exam_banks (semester, exam_type, title, file_id, file_type) VALUES (?,?,?,?,?)",
        (sem, exam_type, title, file_id, file_type)
    )
    db.commit()

def db_get_exam_banks(sem, exam_type):
    db = get_db()
    return db.execute(
        "SELECT * FROM exam_banks WHERE semester=? AND exam_type=? ORDER BY id",
        (sem, exam_type)
    ).fetchall()

def db_delete_exam_bank(bank_id):
    db = get_db()
    db.execute("DELETE FROM exam_banks WHERE id=?", (bank_id,))
    db.commit()

def db_search(query):
    import difflib
    query_lower = query.lower().strip()
    db = get_db()

    # أولاً: بحث LIKE عادي (سريع)
    q = f"%{query}%"
    exact = db.execute(
        "SELECT * FROM lectures WHERE title LIKE ? OR content LIKE ? OR subject LIKE ? "
        "ORDER BY semester,module,subject,id",
        (q, q, q)
    ).fetchall()

    # ثانياً: fuzzy على العناوين لو النتايج أقل من 3
    if len(exact) < 3:
        all_lecs = db.execute("SELECT * FROM lectures").fetchall()
        fuzzy = []
        for lec in all_lecs:
            # تجنب التكرار مع نتايج الـ LIKE
            if any(lec["id"] == e["id"] for e in exact):
                continue
            # نحسب نسبة التشابه مع العنوان والمادة
            ratio_title   = difflib.SequenceMatcher(None, query_lower, lec["title"].lower()).ratio()
            ratio_subject = difflib.SequenceMatcher(None, query_lower, lec["subject"].lower()).ratio()
            ratio = max(ratio_title, ratio_subject)
            if ratio >= 0.55:  # حد التشابه 55%
                fuzzy.append((ratio, lec))
        fuzzy.sort(key=lambda x: x[0], reverse=True)
        extra = [lec for _, lec in fuzzy[:5]]
        return list(exact) + extra

    return list(exact)

def db_increment_views(lid):
    with get_db() as db:
        db.execute(
            "INSERT INTO analytics (lecture_id,views) VALUES (?,1) "
            "ON CONFLICT(lecture_id) DO UPDATE SET views=views+1", (lid,)
        )

def db_get_views(lid):
    with get_db() as db:
        r = db.execute("SELECT views FROM analytics WHERE lecture_id=?", (lid,)).fetchone()
        return r["views"] if r else 0

def db_toggle_bookmark(uid, lid):
    with get_db() as db:
        ex = db.execute(
            "SELECT 1 FROM bookmarks WHERE user_id=? AND lecture_id=?", (uid, lid)
        ).fetchone()
        if ex:
            db.execute("DELETE FROM bookmarks WHERE user_id=? AND lecture_id=?", (uid, lid))
            return False
        db.execute("INSERT INTO bookmarks (user_id,lecture_id) VALUES (?,?)", (uid, lid))
        return True

def db_get_bookmarks(uid):
    with get_db() as db:
        return db.execute(
            "SELECT l.* FROM lectures l JOIN bookmarks b ON l.id=b.lecture_id WHERE b.user_id=?",
            (uid,)
        ).fetchall()

def db_add_quiz(lid, question, opts, answer):
    while len(opts) < 4:
        opts.append("")
    with get_db() as db:
        db.execute(
            "INSERT INTO quizzes (lecture_id,question,opt_a,opt_b,opt_c,opt_d,answer) "
            "VALUES (?,?,?,?,?,?,?)",
            (lid, question, opts[0], opts[1], opts[2], opts[3], answer)
        )

def db_get_quizzes(lid):
    with get_db() as db:
        return db.execute("SELECT * FROM quizzes WHERE lecture_id=?", (lid,)).fetchall()

def db_schedule_broadcast(msg, send_at):
    with get_db() as db:
        db.execute("INSERT INTO scheduled_broadcasts (message,send_at) VALUES (?,?)", (msg, send_at))

def db_pending_broadcasts():
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    with get_db() as db:
        return db.execute(
            "SELECT * FROM scheduled_broadcasts WHERE sent=0 AND send_at<=?", (now,)
        ).fetchall()

def db_mark_sent(bid):
    with get_db() as db:
        db.execute("UPDATE scheduled_broadcasts SET sent=1 WHERE id=?", (bid,))

def db_add_module(sem, module):
    with get_db() as db:
        db.execute(
            "INSERT OR IGNORE INTO custom_modules (semester, module) VALUES (?,?)",
            (sem, module)
        )

def db_add_subject(sem, module, subject):
    with get_db() as db:
        db.execute(
            "INSERT OR IGNORE INTO custom_subjects (semester, module, subject) VALUES (?,?,?)",
            (sem, module, subject)
        )

def db_rename_subject(sem, module, old_subject, new_subject):
    """Rename a subject in custom_subjects and update all lectures."""
    with get_db() as db:
        db.execute(
            "UPDATE custom_subjects SET subject=? WHERE semester=? AND module=? AND subject=?",
            (new_subject, sem, module, old_subject)
        )
        db.execute(
            "UPDATE lectures SET subject=? WHERE semester=? AND module=? AND subject=?",
            (new_subject, sem, module, old_subject)
        )

def db_delete_subject(sem, module, subject):
    """Delete a subject and all its lectures (and related data).
    Works for both hardcoded DATA subjects and custom DB subjects."""
    db = get_db()
    lids = [r[0] for r in db.execute(
        "SELECT id FROM lectures WHERE semester=? AND module=? AND subject=?",
        (sem, module, subject)
    ).fetchall()]
    for lid in lids:
        db.execute("DELETE FROM quizzes   WHERE lecture_id=?", (lid,))
        db.execute("DELETE FROM bookmarks WHERE lecture_id=?", (lid,))
        db.execute("DELETE FROM analytics WHERE lecture_id=?", (lid,))
    db.execute(
        "DELETE FROM lectures WHERE semester=? AND module=? AND subject=?",
        (sem, module, subject)
    )
    db.execute(
        "DELETE FROM custom_subjects WHERE semester=? AND module=? AND subject=?",
        (sem, module, subject)
    )
    # لو المادة hardcoded في DATA، نضيفها في جدول خاص عشان نعرف إنها اتحذفت
    hardcoded_subs = DATA.get(sem, {}).get(module, [])
    if subject in hardcoded_subs:
        db.execute(
            "CREATE TABLE IF NOT EXISTS deleted_hardcoded_subjects "
            "(semester TEXT, module TEXT, subject TEXT, PRIMARY KEY(semester,module,subject))"
        )
        db.execute(
            "INSERT OR IGNORE INTO deleted_hardcoded_subjects (semester,module,subject) VALUES (?,?,?)",
            (sem, module, subject)
        )
    db.commit()

def db_delete_module(sem, module):
    """Delete a module, all its subjects and lectures."""
    db = get_db()
    lids = [r[0] for r in db.execute(
        "SELECT id FROM lectures WHERE semester=? AND module=?",
        (sem, module)
    ).fetchall()]
    for lid in lids:
        db.execute("DELETE FROM quizzes   WHERE lecture_id=?", (lid,))
        db.execute("DELETE FROM bookmarks WHERE lecture_id=?", (lid,))
        db.execute("DELETE FROM analytics WHERE lecture_id=?", (lid,))
    db.execute("DELETE FROM lectures       WHERE semester=? AND module=?", (sem, module))
    db.execute("DELETE FROM custom_subjects WHERE semester=? AND module=?", (sem, module))
    db.execute("DELETE FROM custom_modules  WHERE semester=? AND module=?", (sem, module))
    db.commit()

def db_rename_module(sem, old_module, new_module):
    """Rename a module in custom_modules and update all lectures/subjects."""
    with get_db() as db:
        db.execute(
            "UPDATE custom_modules SET module=? WHERE semester=? AND module=?",
            (new_module, sem, old_module)
        )
        db.execute(
            "UPDATE custom_subjects SET module=? WHERE semester=? AND module=?",
            (new_module, sem, old_module)
        )
        db.execute(
            "UPDATE lectures SET module=? WHERE semester=? AND module=?",
            (new_module, sem, old_module)
        )

def db_get_modules(sem):
    """Return merged list: hardcoded DATA modules + custom DB modules."""
    base = list(DATA.get(sem, {}).keys())
    with get_db() as db:
        rows = db.execute(
            "SELECT module FROM custom_modules WHERE semester=? ORDER BY id",
            (sem,)
        ).fetchall()
    extra = [r["module"] for r in rows if r["module"] not in base]
    return base + extra

def db_get_subjects(sem, module):
    """Return merged list: hardcoded DATA subjects + custom DB subjects,
    excluding any hardcoded subjects that were explicitly deleted."""
    base = list(DATA.get(sem, {}).get(module, []))
    db = get_db()
    # استثني المواد الـ hardcoded اللي اتحذفت
    try:
        deleted = {r[0] for r in db.execute(
            "SELECT subject FROM deleted_hardcoded_subjects WHERE semester=? AND module=?",
            (sem, module)
        ).fetchall()}
    except Exception:
        deleted = set()
    base = [s for s in base if s not in deleted]
    rows = db.execute(
        "SELECT subject FROM custom_subjects WHERE semester=? AND module=? ORDER BY id",
        (sem, module)
    ).fetchall()
    extra = [r["subject"] for r in rows if r["subject"] not in base]
    return base + extra

def db_stats():
    with get_db() as db:
        u = db.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        l = db.execute("SELECT COUNT(*) FROM lectures").fetchone()[0]
        top = db.execute(
            "SELECT l.title,l.module,l.subject,a.views FROM analytics a "
            "JOIN lectures l ON l.id=a.lecture_id ORDER BY a.views DESC LIMIT 5"
        ).fetchall()
        by_mod = db.execute(
            "SELECT module,COUNT(*) cnt FROM lectures GROUP BY module"
        ).fetchall()
        all_users = db.execute(
            "SELECT user_id, name, username, joined_at FROM users ORDER BY joined_at DESC"
        ).fetchall()
    return u, l, top, by_mod, all_users

# ─────────────────────────────────────────────────────
# 🔒  RATE LIMITING
# ─────────────────────────────────────────────────────
_rate: dict = defaultdict(list)

def is_rate_limited(uid):
    if uid in ADMIN_IDS:
        return False
    now = asyncio.get_event_loop().time()
    _rate[uid] = [t for t in _rate[uid] if now - t < 10]
    if len(_rate[uid]) >= 10:
        return True
    _rate[uid].append(now)
    return False

# ─────────────────────────────────────────────────────
# 🔧  HELPERS
# ─────────────────────────────────────────────────────
def is_admin(uid):
    if uid in ADMIN_IDS:
        return True
    db = get_db()
    return db.execute("SELECT 1 FROM admins WHERE user_id=?", (uid,)).fetchone() is not None

def db_add_admin(uid, name):
    db = get_db()
    db.execute("INSERT OR IGNORE INTO admins (user_id, name) VALUES (?,?)", (uid, name))
    db.commit()

def db_remove_admin(uid):
    db = get_db()
    db.execute("DELETE FROM admins WHERE user_id=?", (uid,))
    db.commit()

def db_list_admins():
    db = get_db()
    return db.execute("SELECT user_id, name, added_at FROM admins ORDER BY added_at").fetchall()
def file_icon(ft):   return {"video":"🎬","photo":"🖼","document":"📎","pdf":"📕"}.get(ft,"📄")

def all_modules():
    result = []
    for sem in DATA:
        for m in db_get_modules(sem):
            if m not in result:
                result.append(m)
    return result

def find_sem_for_module(module):
    for s in DATA:
        if module in db_get_modules(s):
            return s
    return None

def all_subjects():
    result = []
    for sem in DATA:
        for mod in db_get_modules(sem):
            for sub in db_get_subjects(sem, mod):
                result.append(sub)
    return result

# ─────────────────────────────────────────────────────
# 🎛️  KEYBOARDS  — Student
# ─────────────────────────────────────────────────────
def kb_semesters():
    sem_icons = {"Sem 1": "1️⃣", "Sem 2": "2️⃣", "Sem 3": "3️⃣", "Sem 4": "4️⃣"}
    rows = []
    sem_list = list(DATA.keys())
    # Two semesters per row
    for i in range(0, len(sem_list), 2):
        row = []
        for s in sem_list[i:i+2]:
            icon = sem_icons.get(s, "📘")
            row.append(InlineKeyboardButton(f"{icon} {s}", callback_data=f"sem:{s}"))
        rows.append(row)
    rows.append([
        InlineKeyboardButton("🔖 محفوظاتي", callback_data="bookmarks"),
        InlineKeyboardButton("🔍 بحث",      callback_data="search"),
        InlineKeyboardButton("👤 بروفايلي", callback_data="profile"),
    ])
    return InlineKeyboardMarkup(rows)

def kb_modules(sem):
    modules = db_get_modules(sem)
    counts  = db_module_counts_batch(sem)
    rows = []
    for i in range(0, len(modules), 2):
        row = []
        for module in modules[i:i+2]:
            cnt   = counts.get(module, 0)
            short = module if len(module) <= 22 else module[:20] + "…"
            row.append(InlineKeyboardButton(
                f"📗 {short}  ({cnt})", callback_data=f"mod:{sem}:{module}"
            ))
        rows.append(row)
    rows.append([
        InlineKeyboardButton("🏦 بنك الميدتيرم", callback_data=f"bank:{sem}:midterm"),
        InlineKeyboardButton("🏦 بنك الفاينال",  callback_data=f"bank:{sem}:final"),
    ])
    rows.append([InlineKeyboardButton("🔙 رجوع", callback_data="back:semesters")])
    return InlineKeyboardMarkup(rows)

CONTENT_TYPES = [
    ("📖 Lectures",       "L"),
    ("🔬 Practicals",     "P"),
    ("❓ Question Banks", "Q"),
    ("📝 Summaries",      "S"),
]

CTYPE_FULL = {
    "L": "lectures",
    "P": "practicals",
    "Q": "question_banks",
    "S": "summaries",
}

def ctype_label(code):
    names = {"L":"📖 Lectures","P":"🔬 Practicals","Q":"❓ Question Banks","S":"📝 Summaries"}
    return names.get(code, code)

def ctype_db(code):
    return CTYPE_FULL.get(code, "lectures")

def kb_exam_bank(sem, exam_type):
    banks = db_get_exam_banks(sem, exam_type)
    label = "الميدتيرم" if exam_type == "midterm" else "الفاينال"
    rows = []
    for b in banks:
        rows.append([InlineKeyboardButton(
            f"📄 {b['title']}", callback_data=f"bankfile:{b['id']}"
        )])
    if not rows:
        rows.append([InlineKeyboardButton("📭 مفيش ملفات لسه", callback_data="noop")])
    rows.append([InlineKeyboardButton("🔙 رجوع", callback_data=f"sem:{sem}")])
    return InlineKeyboardMarkup(rows)

def kb_content_types(sem, module):
    rows = [
        [InlineKeyboardButton("📖 محاضرات",     callback_data=f"ctype:{sem}:{module}:L"),
         InlineKeyboardButton("🔬 عملي",        callback_data=f"ctype:{sem}:{module}:P")],
        [InlineKeyboardButton("❓ بنك أسئلة",   callback_data=f"ctype:{sem}:{module}:Q"),
         InlineKeyboardButton("📝 ملخصات",      callback_data=f"ctype:{sem}:{module}:S")],
        [InlineKeyboardButton("🔙 رجوع", callback_data=f"back:modules:{sem}")],
    ]
    return InlineKeyboardMarkup(rows)

def kb_subjects(sem, module, ctype="L"):
    subjects = db_get_subjects(sem, module)
    counts   = db_subject_counts_batch(sem, module, ctype_db(ctype))
    rows = []
    for subject in subjects:
        cnt   = counts.get(subject, 0)
        nid   = nav_save(sem, module, ctype, subject)
        label = f"📄 {subject}  ({cnt})" if cnt else f"📄 {subject}  —"
        rows.append([InlineKeyboardButton(label, callback_data=f"sub:{nid}:0")])
    rows.append([InlineKeyboardButton("🔙 رجوع", callback_data=f"ctype_back:{sem}:{module}")])
    return InlineKeyboardMarkup(rows)

def kb_lectures(sem, module, subject, page=0, ctype="L"):
    rows_db, total = db_get_lectures(sem, module, subject, page, ctype_db(ctype))
    nid = nav_save(sem, module, ctype, subject)
    rows = []
    for lec in rows_db:
        icon = file_icon(lec["file_type"])
        rows.append([InlineKeyboardButton(
            f"{icon} {lec['title']}",
            callback_data=f"lec:{lec['id']}:{ctype}"
        )])
    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("◀️ السابق", callback_data=f"sub:{nid}:{page-1}"))
    if (page + 1) * PAGE_SIZE < total:
        nav.append(InlineKeyboardButton("التالي ▶️", callback_data=f"sub:{nid}:{page+1}"))
    if nav:
        rows.append(nav)
    rows.append([InlineKeyboardButton("🔙 رجوع", callback_data=f"ctype_back:{sem}:{module}")])
    return InlineKeyboardMarkup(rows), total

def kb_lecture_view(sem, module, subject, lid, uid, ctype="L"):
    with get_db() as db:
        bm = bool(db.execute(
            "SELECT 1 FROM bookmarks WHERE user_id=? AND lecture_id=?", (uid, lid)
        ).fetchone())
    quizzes = db_get_quizzes(lid)
    nid = nav_save(sem, module, ctype, subject)
    rows = [[InlineKeyboardButton(
        "🔖 محفوظ ✓" if bm else "🔖 احفظ",
        callback_data=f"bm:{lid}:{ctype}"
    )]]
    if quizzes:
        rows.append([InlineKeyboardButton(f"🧠 اختبر نفسك  ({len(quizzes)} سؤال)", callback_data=f"quiz:{lid}:0")])
    rows.append([InlineKeyboardButton("🔙 رجوع للقائمة", callback_data=f"sub:{nid}:0")])
    return InlineKeyboardMarkup(rows)

def kb_quiz(quiz, idx):
    rows = []
    for letter, opt_key in [("A","opt_a"),("B","opt_b"),("C","opt_c"),("D","opt_d")]:
        if quiz[opt_key]:
            rows.append([InlineKeyboardButton(
                f"{letter}. {quiz[opt_key]}",
                callback_data=f"qa:{quiz['lecture_id']}:{idx}:{letter}"
            )])
    return InlineKeyboardMarkup(rows)

# ─────────────────────────────────────────────────────
# 🎛️  KEYBOARDS  — Admin
# ─────────────────────────────────────────────────────
def admin_kb():
    return ReplyKeyboardMarkup(
        [["➕ Add Content",      "✏️ Edit Lecture"],
         ["🗑 Delete Lecture",   "🧠 Add Quiz"],
         ["📦 Add Module",       "📂 Add Subject"],
         ["✏️ Edit Module",      "🗑 Delete Module"],
         ["✏️ Edit Subject",     "🗑 Delete Subject"],
         ["📢 Broadcast",       "📨 رسالة ليوزر"],
         ["📊 Stats",           "💾 Backup"],
         ["📥 Import Excel",    "📤 Export Excel"],
         ["📦 Import ZIP",      "🔎 بحث يوزر"],
         ["🏦 إدارة البنوك",    "🔤 ترتيب أبجدي"],
         ["👑 قائمة الأدمنز",  "👑 إضافة أدمن",  "🗑 حذف أدمن"],
         ["🔙 Exit Admin"]],
        resize_keyboard=True
    )

def all_sems():
    """Return all semesters: hardcoded + any extra in DB."""
    base = list(DATA.keys())
    db   = get_db()
    extras = set()
    for tbl in ("custom_modules", "custom_subjects", "lectures"):
        for r in db.execute(f"SELECT DISTINCT semester FROM {tbl}").fetchall():
            if r[0] and r[0] not in base:
                extras.add(r[0])
    return base + sorted(extras)

def kb_admin_sems(prefix):
    rows = [[InlineKeyboardButton(f"📘 {s}", callback_data=f"{prefix}_sem:{s}")]
            for s in all_sems()]
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
    return InlineKeyboardMarkup(rows)

def kb_admin_modules(prefix, sem):
    rows = []
    for module in db_get_modules(sem):
        cnt = db_module_count(sem, module)
        rows.append([InlineKeyboardButton(
            f"📗 {module}  ({cnt})", callback_data=f"{prefix}_mod:{sem}:{module}"
        )])
    rows.append([InlineKeyboardButton("🔙 Back", callback_data=f"{prefix}_sem_back")])
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
    return InlineKeyboardMarkup(rows)

def kb_admin_subjects(prefix, sem, module):
    rows = []
    for subject in db_get_subjects(sem, module):
        cnt = db_subject_count(sem, module, subject)
        rows.append([InlineKeyboardButton(
            f"📄 {subject}  ({cnt})", callback_data=f"{prefix}_sub:{sem}:{module}:{subject}"
        )])
    rows.append([InlineKeyboardButton("🔙 Back", callback_data=f"{prefix}_mod_back:{sem}")])
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
    return InlineKeyboardMarkup(rows)

def kb_admin_lectures(prefix, sem, module, subject):
    with get_db() as db:
        lecs = db.execute(
            "SELECT * FROM lectures WHERE semester=? AND module=? AND subject=? ORDER BY id",
            (sem, module, subject)
        ).fetchall()
    lecs = sorted(lecs, key=_natural_key)
    rows = [[InlineKeyboardButton(
        f"{lec['title']}", callback_data=f"{prefix}_lec:{lec['id']}"
    )] for lec in lecs]
    rows.append([InlineKeyboardButton("🔙 Back", callback_data=f"{prefix}_sub_back:{sem}:{module}")])
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
    return InlineKeyboardMarkup(rows)

def kb_edit_fields(lid):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📝 Title",   callback_data=f"edf:{lid}:title")],
        [InlineKeyboardButton("📄 Content", callback_data=f"edf:{lid}:content")],
        [InlineKeyboardButton("📎 File",    callback_data=f"edf:{lid}:file")],
        [InlineKeyboardButton("❌ Cancel",  callback_data="admin:cancel")],
    ])

def kb_sems_for_new_module():
    rows = [[InlineKeyboardButton(f"📘 {s}", callback_data=f"newmod_sem:{s}")] for s in all_sems()]
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
    return InlineKeyboardMarkup(rows)

def kb_sems_for_new_subject():
    rows = [[InlineKeyboardButton(f"📘 {s}", callback_data=f"newsub_sem:{s}")] for s in all_sems()]
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
    return InlineKeyboardMarkup(rows)

def kb_sems_for_edit_module():
    rows = [[InlineKeyboardButton(f"📘 {s}", callback_data=f"editmod_sem:{s}")] for s in all_sems()]
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
    return InlineKeyboardMarkup(rows)

def kb_sems_for_del_module():
    sems = list(DATA.keys())
    db   = get_db()
    db_sems = [r[0] for r in db.execute(
        "SELECT DISTINCT semester FROM custom_modules"
    ).fetchall() if r[0] not in sems]
    rows = [[InlineKeyboardButton(f"📘 {s}", callback_data=f"delmod_sem:{s}")]
            for s in sems + db_sems]
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
    return InlineKeyboardMarkup(rows)

def kb_mods_for_edit_module(sem):
    rows = [[InlineKeyboardButton(f"📗 {m}", callback_data=f"editmod_pick:{sem}:{m}")]
            for m in db_get_modules(sem)]
    rows.append([InlineKeyboardButton("🔙 Back", callback_data="editmod_back_sem")])
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
    return InlineKeyboardMarkup(rows)

def kb_mods_for_del_module(sem):
    rows = [[InlineKeyboardButton(f"📗 {m}", callback_data=f"delmod_pick:{sem}:{m}")]
            for m in db_get_modules(sem)]
    rows.append([InlineKeyboardButton("🔙 Back", callback_data="delmod_back_sem")])
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
    return InlineKeyboardMarkup(rows)

def kb_sems_for_edit_subject():
    rows = [[InlineKeyboardButton(f"📘 {s}", callback_data=f"editsub_sem:{s}")] for s in all_sems()]
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
    return InlineKeyboardMarkup(rows)

def kb_sems_for_del_subject():
    sems = list(DATA.keys())
    db   = get_db()
    db_sems = [r[0] for r in db.execute(
        "SELECT DISTINCT semester FROM custom_subjects"
    ).fetchall() if r[0] not in sems]
    rows = [[InlineKeyboardButton(f"📘 {s}", callback_data=f"delsub_sem:{s}")]
            for s in sems + db_sems]
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
    return InlineKeyboardMarkup(rows)

def kb_mods_for_edit_subject(sem):
    rows = [[InlineKeyboardButton(f"📗 {m}", callback_data=f"editsub_mod:{sem}:{m}")]
            for m in db_get_modules(sem)]
    rows.append([InlineKeyboardButton("🔙 Back", callback_data="editsub_back_sem")])
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
    return InlineKeyboardMarkup(rows)

def kb_mods_for_del_subject(sem):
    rows = [[InlineKeyboardButton(f"📗 {m}", callback_data=f"delsub_mod:{sem}:{m}")]
            for m in db_get_modules(sem)]
    rows.append([InlineKeyboardButton("🔙 Back", callback_data="delsub_back_sem")])
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
    return InlineKeyboardMarkup(rows)

def kb_subs_for_edit(sem, module):
    rows = [[InlineKeyboardButton(f"📄 {s}", callback_data=f"editsub_pick:{sem}:{module}:{s}")]
            for s in db_get_subjects(sem, module)]
    rows.append([InlineKeyboardButton("🔙 Back", callback_data=f"editsub_back_mod:{sem}")])
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
    return InlineKeyboardMarkup(rows)

def kb_subs_for_del(sem, module):
    rows = [[InlineKeyboardButton(f"📄 {s}", callback_data=f"delsub_pick:{sem}:{module}:{s}")]
            for s in db_get_subjects(sem, module)]
    rows.append([InlineKeyboardButton("🔙 Back", callback_data=f"delsub_back_mod:{sem}")])
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
    return InlineKeyboardMarkup(rows)

def kb_mods_for_new_subject(sem):
    rows = [[InlineKeyboardButton(f"📗 {m}", callback_data=f"newsub_mod:{sem}:{m}")]
            for m in db_get_modules(sem)]
    rows.append([InlineKeyboardButton("🔙 Back", callback_data="newsub_back_sem")])
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
    return InlineKeyboardMarkup(rows)

# ─────────────────────────────────────────────────────
# 🚀  /start
# ─────────────────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    is_new = db_register(update.effective_user)
    context.user_data.clear()
    name = md_escape(update.effective_user.first_name or "طالب")
    text = (
        f"👋 أهلاً *{name}!*\n\n"
        f"🩺 *DU - Medicine 2024* — رفيقك الطبي الأول\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"📖 محاضرات  •  🔬 عملي  •  ❓ بنك أسئلة  •  📝 ملخصات\n"
        f"━━━━━━━━━━━━━━━━━━━━\n\n"
        f"اختار ترمك وابدأ 👇"
    )
    if update.message:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb_semesters())
    else:
        await update.callback_query.edit_message_text(text, parse_mode="Markdown", reply_markup=kb_semesters())

    # إشعار الأدمن بيوزر جديد
    if is_new:
        db  = get_db()
        total = db.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        uname = f"@{update.effective_user.username}" if update.effective_user.username else "—"
        notif = (
            f"🔔 <b>يوزر جديد انضم!</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"📛 الاسم: {_html(update.effective_user.full_name)}\n"
            f"🔗 يوزرنيم: {_html(uname)}\n"
            f"🆔 ID: <code>{update.effective_user.id}</code>\n"
            f"👥 إجمالي المستخدمين: <b>{total}</b>"
        )
        for aid in ADMIN_IDS:
            try:
                await context.bot.send_message(aid, notif, parse_mode="HTML")
            except: pass

# ─────────────────────────────────────────────────────
# 👑  /admin
# ─────────────────────────────────────────────────────
async def admin_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("❌ Not authorized.")
        return
    context.user_data.clear()
    u, l, _, _, _ = db_stats()
    await update.message.reply_text(
        f"🛠 *لوحة التحكم*\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"👥 المستخدمين: `{u}`  •  📚 المحتوى: `{l}`\n"
        f"━━━━━━━━━━━━━━━━━━━━\n\n"
        f"اختار العملية 👇",
        parse_mode="Markdown", reply_markup=admin_kb()
    )

# ─────────────────────────────────────────────────────
# 🔍  /search
# ─────────────────────────────────────────────────────
async def search_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("🔍 Usage: `/search keyword`", parse_mode="Markdown")
        return
    _do_search(update, context, " ".join(context.args))

async def _do_search(update, context, query):
    results = db_search(query)
    if not results:
        msg = f"😕 *مفيش نتايج لـ* _{md_escape(query)}_\n\nجرب كلمة تانية 🔍"
        if update.message:
            await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=kb_semesters())
        else:
            await update.callback_query.edit_message_text(msg, parse_mode="Markdown",
                                                          reply_markup=InlineKeyboardMarkup([[
                                                              InlineKeyboardButton("🔙 رجوع", callback_data="back:semesters")
                                                          ]]))
        return
    rows = []
    for lec in results[:10]:
        rows.append([InlineKeyboardButton(
            f"📄 {lec['title']} — {lec['subject']}",
            callback_data=f"lec:{lec['id']}:L"
        )])
    rows.append([InlineKeyboardButton("🔙 رجوع", callback_data="back:semesters")])
    kb = InlineKeyboardMarkup(rows)
    text = f"🔍 *{len(results)} نتيجة* لـ _{md_escape(query)}_\n━━━━━━━━━━━━━━━━━━━━"
    if update.message:
        await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)
    else:
        await update.callback_query.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)

# ─────────────────────────────────────────────────────
# 🔘  CALLBACK HANDLER
# ─────────────────────────────────────────────────────
def _html(text: str) -> str:
    """Escape special HTML characters to prevent parse errors."""
    if not text:
        return text
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

async def _send_user_result(update, context, p, search_uid):
    """Helper: بيبعت نتيجة يوزر واحد."""
    u      = p["user"]
    name   = _html(u["name"] or "—")
    uname  = _html(f"@{u['username']}" if u["username"] else "—")
    joined = (u["joined_at"] or "")[:10]
    last   = (u["last_seen"] or "")[:16]
    avg    = p["avg_pct"]
    lines  = [
        f"🔎 <b>نتيجة البحث</b>",
        f"━━━━━━━━━━━━━━━━━━━━",
        f"🆔 ID: <code>{u['user_id']}</code>",
        f"📛 الاسم: {name}",
        f"🔗 يوزرنيم: {uname}",
        f"📅 انضم: {joined}",
        f"🕐 آخر ظهور: {last}",
        f"━━━━━━━━━━━━━━━━━━━━",
        f"🧠 محاولات الكويز: {p['quiz_attempts']}",
        f"📊 متوسط الدرجات: {avg}%",
        f"🔖 المحفوظات: {p['bookmarks']}",
    ]
    if p["recent_quizzes"]:
        lines.append("\n📝 آخر 5 كويزات:")
        for q in p["recent_quizzes"]:
            pct   = round(q["score"] / q["total"] * 100) if q["total"] else 0
            date  = (q["taken_at"] or "")[:10]
            title = _html(q['title'][:25])
            lines.append(f"  • {title} — {pct}% ({date})")
    await update.message.reply_text(
        "\n".join(lines), parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("📄 تقرير Excel", callback_data=f"report_uid:{search_uid}"),
            InlineKeyboardButton("📋 تقرير PDF",   callback_data=f"report_pdf:{search_uid}"),
        ]])
    )
    await update.message.reply_text("👆", reply_markup=admin_kb())

async def callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q   = update.callback_query
    await q.answer()
    data = q.data
    uid  = update.effective_user.id

    if is_rate_limited(uid):
        await q.answer("⏳ شوية شوية! متضغطش بسرعة 😅", show_alert=True)
        return

    # ════════════════════════════════════════════════
    # STUDENT NAVIGATION
    # ════════════════════════════════════════════════

    # Semester → Modules
    if data.startswith("sem:"):
        sem = data[4:]
        await q.edit_message_text(
            f"📘 *{sem}*\nاختار الموديول اللي عايزه 👇",
            parse_mode="Markdown", reply_markup=kb_modules(sem))

    elif data.startswith("bank:"):
        _, sem, exam_type = data.split(":", 2)
        label = "🏦 بنك الميدتيرم" if exam_type == "midterm" else "🏦 بنك الفاينال"
        await q.edit_message_text(
            f"📘 *{sem}*  ›  *{label}*\n\nاختار الملف اللي عايزه 👇",
            parse_mode="Markdown", reply_markup=kb_exam_bank(sem, exam_type)
        )

    elif data.startswith("bankfile:"):
        bank_id = int(data.split(":")[1])
        db = get_db()
        b = db.execute("SELECT * FROM exam_banks WHERE id=?", (bank_id,)).fetchone()
        if not b:
            await q.answer("❌ الملف مش موجود!", show_alert=True)
            return
        await q.answer()
        send = {
            "document": context.bot.send_document,
            "photo":    context.bot.send_photo,
            "video":    context.bot.send_video,
        }.get(b["file_type"], context.bot.send_document)
        await send(chat_id=q.message.chat_id, **{b["file_type"]: b["file_id"]}, caption=f"📄 *{b['title']}*", parse_mode="Markdown")
        return

    elif data == "noop":
        await q.answer()
        return

    # ── Admin Bank Management ──────────────────────────
    elif data.startswith("adminbank_sem:"):
        sem = data.split(":", 1)[1]
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("🏦 ميدتيرم", callback_data=f"adminbank_type:{sem}:midterm"),
             InlineKeyboardButton("🏦 فاينال",  callback_data=f"adminbank_type:{sem}:final")],
            [InlineKeyboardButton("🔙 رجوع", callback_data="admin:cancel")]
        ])
        await q.edit_message_text(f"🏦 *بنوك {sem}*\nاختار النوع:", parse_mode="Markdown", reply_markup=kb)

    elif data.startswith("adminbank_type:"):
        _, sem, exam_type = data.split(":", 2)
        label = "الميدتيرم" if exam_type == "midterm" else "الفاينال"
        banks = db_get_exam_banks(sem, exam_type)
        rows = [[InlineKeyboardButton(
            f"🗑 {b['title']}", callback_data=f"adminbank_del:{b['id']}:{sem}:{exam_type}"
        )] for b in banks]
        rows.append([InlineKeyboardButton(
            "➕ رفع ملف جديد", callback_data=f"adminbank_upload:{sem}:{exam_type}"
        )])
        rows.append([InlineKeyboardButton("🔙 رجوع", callback_data=f"adminbank_sem:{sem}")])
        txt = f"🏦 *بنك {label} — {sem}*\n{'اضغط على الملف لحذفه' if banks else '📭 مفيش ملفات لسه'}"
        await q.edit_message_text(txt, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(rows))

    elif data.startswith("adminbank_del:"):
        parts = data.split(":")
        bank_id, sem, exam_type = int(parts[1]), parts[2], parts[3]
        db_delete_exam_bank(bank_id)
        await q.answer("✅ اتحذف!")
        # refresh
        label = "الميدتيرم" if exam_type == "midterm" else "الفاينال"
        banks = db_get_exam_banks(sem, exam_type)
        rows = [[InlineKeyboardButton(
            f"🗑 {b['title']}", callback_data=f"adminbank_del:{b['id']}:{sem}:{exam_type}"
        )] for b in banks]
        rows.append([InlineKeyboardButton("➕ رفع ملف جديد", callback_data=f"adminbank_upload:{sem}:{exam_type}")])
        rows.append([InlineKeyboardButton("🔙 رجوع", callback_data=f"adminbank_sem:{sem}")])
        txt = f"🏦 *بنك {label} — {sem}*\n{'اضغط على الملف لحذفه' if banks else '📭 مفيش ملفات لسه'}"
        await q.edit_message_text(txt, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(rows))

    elif data.startswith("adminbank_upload:"):
        _, sem, exam_type = data.split(":", 2)
        label = "الميدتيرم" if exam_type == "midterm" else "الفاينال"
        context.user_data["step"] = "adminbank_upload"
        context.user_data["bank_sem"] = sem
        context.user_data["bank_type"] = exam_type
        await q.edit_message_text(
            f"📤 *رفع ملف — بنك {label} ({sem})*\n\nابعت الملف دلوقتي (PDF, صورة, أو فيديو)\n"
            f"أو ابعت /cancel للإلغاء",
            parse_mode="Markdown"
        )

    # Module → Content Types
    elif data.startswith("mod:"):
        _, sem, module = data.split(":", 2)
        await q.edit_message_text(
            f"📘 {sem}  ›  📗 *{module}*\n\nاختار نوع المحتوى 👇",
            parse_mode="Markdown", reply_markup=kb_content_types(sem, module)
        )

    # Content Type → Subjects
    elif data.startswith("ctype:"):
        parts  = data.split(":", 3)
        sem    = parts[1]
        module = parts[2]
        ctype  = parts[3]
        await q.edit_message_text(
            f"📘 {sem}  ›  📗 {module}  ›  {ctype_label(ctype)}\n\nاختار المادة 👇",
            parse_mode="Markdown", reply_markup=kb_subjects(sem, module, ctype)
        )

    # Back to content types
    elif data.startswith("ctype_back:"):
        _, sem, module = data.split(":", 2)
        await q.edit_message_text(
            f"📘 {sem}  ›  📗 *{module}*\n\nاختار نوع المحتوى 👇",
            parse_mode="Markdown", reply_markup=kb_content_types(sem, module)
        )

    # Subject → Lectures (with pagination)
    elif data.startswith("sub:"):
        parts = data.split(":", 2)
        nid   = int(parts[1])
        page  = int(parts[2]) if len(parts) > 2 else 0
        sem, module, ctype, subject = nav_load(nid)
        if not sem:
            await q.edit_message_text("⚠️ Session expired. Please /start again.")
            return
        kb, total = kb_lectures(sem, module, subject, page, ctype)
        total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
        label = ctype_label(ctype)
        if total == 0:
            await q.edit_message_text(
                f"📭 *{subject}*\n\nلا يوجد {label} متاح حالياً.\nترقب الإضافات قريباً! 🔜",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔙 رجوع", callback_data=f"ctype_back:{sem}:{module}")
                ]])
            )
            return
        await q.edit_message_text(
            f"📘 {sem}  ›  {label}\n"
            f"📄 *{subject}*\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"الصفحة {page+1}/{total_pages}  •  {total} عنصر",
            parse_mode="Markdown", reply_markup=kb
        )

    # Lecture content
    elif data.startswith("lec:"):
        parts   = data.split(":", 2)
        lid     = int(parts[1])
        ctype   = parts[2] if len(parts) > 2 else "L"
        lec     = db_get_lecture(lid)
        if not lec:
            await q.edit_message_text("⚠️ Lecture not found.")
            return
        sem     = lec["semester"]
        module  = lec["module"]
        subject = lec["subject"]
        db_increment_views(lid)
        views  = db_get_views(lid)
        ctype_icon = {"lectures":"📖","practicals":"🔬","question_banks":"❓","summaries":"📝"}.get(lec["content_type"],"📄")
        header = (
            f"{ctype_icon} *{md_escape(lec['title'])}*\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"📚 _{md_escape(subject)}_  •  _{md_escape(module)}_\n"
            f"👁 {views} مشاهدة\n"
            f"━━━━━━━━━━━━━━━━━━━━\n\n"
            f"{md_escape(lec['content'])}"
        )
        kb_view = kb_lecture_view(sem, module, subject, lid, uid, ctype)
        if lec["file_id"]:
            await q.message.reply_text(header, parse_mode="Markdown", reply_markup=kb_view)
            ft = lec["file_type"]
            if ft == "photo":    await q.message.reply_photo(lec["file_id"])
            elif ft == "video":  await q.message.reply_video(lec["file_id"])
            else:                await q.message.reply_document(lec["file_id"])
            try: await q.message.delete()
            except: pass
        else:
            await q.edit_message_text(header, parse_mode="Markdown", reply_markup=kb_view)

    # Bookmark toggle
    elif data.startswith("bm:"):
        parts = data.split(":", 2)
        lid   = int(parts[1])
        ctype = parts[2] if len(parts) > 2 else "L"
        lec   = db_get_lecture(lid)
        added = db_toggle_bookmark(uid, lid)
        await q.answer("🔖 تمت الإضافة للمحفوظات!" if added else "🗑 تم الحذف من المحفوظات.")
        if lec:
            try: await q.edit_message_reply_markup(
                reply_markup=kb_lecture_view(lec["semester"], lec["module"], lec["subject"], lid, uid, ctype))
            except: pass

    # Bookmarks list
    elif data == "bookmarks":
        bms = db_get_bookmarks(uid)
        if not bms:
            await q.edit_message_text(
                "🔖 *محفوظاتك فاضية!*\n\nافتح أي محاضرة واضغط *احفظ* عشان تلاقيها هنا.",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔙 رجوع", callback_data="back:semesters")
                ]])
            )
            return
        rows = [[InlineKeyboardButton(
            f"🔖 {lec['title']} — {lec['subject']}",
            callback_data=f"lec:{lec['id']}:L"
        )] for lec in bms]
        rows.append([InlineKeyboardButton("🔙 رجوع", callback_data="back:semesters")])
        await q.edit_message_text(
            f"🔖 *محفوظاتك*  ({len(bms)} عنصر)\n━━━━━━━━━━━━━━━━━━━━",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(rows)
        )

    # Admin — view user from search results list
    elif data.startswith("view_user:"):
        target_uid = int(data.split(":")[1])
        p = db_find_user(target_uid)
        if not p:
            await q.answer("❌ مفيش يوزر!")
            return
        await q.answer()
        u      = p["user"]
        uname  = _html(f"@{u['username']}" if u["username"] else "—")
        joined = (u["joined_at"] or "")[:10]
        last   = (u["last_seen"] or "")[:16]
        avg    = p["avg_pct"]
        lines  = [
            f"🔎 <b>بيانات اليوزر</b>",
            f"━━━━━━━━━━━━━━━━━━━━",
            f"🆔 ID: <code>{u['user_id']}</code>",
            f"📛 الاسم: {_html(u['name'] or '—')}",
            f"🔗 يوزرنيم: {uname}",
            f"📅 انضم: <code>{joined}</code>",
            f"🕐 آخر ظهور: <code>{last}</code>",
            f"━━━━━━━━━━━━━━━━━━━━",
            f"🧠 محاولات الكويز: <code>{p['quiz_attempts']}</code>",
            f"📊 متوسط الدرجات: <code>{avg}%</code>",
            f"🔖 المحفوظات: <code>{p['bookmarks']}</code>",
        ]
        await q.edit_message_text(
            "\n".join(lines), parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("📄 تقرير Excel", callback_data=f"report_uid:{target_uid}"),
                InlineKeyboardButton("📋 تقرير PDF",   callback_data=f"report_pdf:{target_uid}"),
            ]])
        )
        return

    # Admin — generate single user Excel report
    elif data.startswith("report_uid:"):
        target_uid = int(data.split(":")[1])
        await q.answer("⏳ بيتجهز التقرير…")
        try:
            xlsx = build_user_report_excel(target_uid)
            stamp = datetime.now().strftime("%Y%m%d_%H%M")
            await context.bot.send_document(
                chat_id=uid,
                document=io.BytesIO(xlsx),
                filename=f"user_{target_uid}_{stamp}.xlsx",
                caption=f"📄 تقرير Excel — يوزر <code>{target_uid}</code>",
                parse_mode="HTML"
            )
        except Exception as e:
            await context.bot.send_message(uid, f"❌ فشل التقرير: {e}")
        return

    # Admin — generate single user PDF report
    elif data.startswith("report_pdf:"):
        target_uid = int(data.split(":")[1])
        await q.answer("⏳ بيتجهز الـ PDF…")
        try:
            pdf = build_user_report_pdf(target_uid)
            stamp = datetime.now().strftime("%Y%m%d_%H%M")
            await context.bot.send_document(
                chat_id=uid,
                document=io.BytesIO(pdf),
                filename=f"user_{target_uid}_{stamp}.pdf",
                caption=f"📋 تقرير PDF — يوزر <code>{target_uid}</code>",
                parse_mode="HTML"
            )
        except Exception as e:
            await context.bot.send_message(uid, f"❌ فشل الـ PDF: {e}")
        return

    # Search inline prompt
    elif data == "search":
        context.user_data["step"] = "searching"
        await q.edit_message_text(
            "🔍 *بحث*\n\nاكتب الكلمة اللي بتدور عليها وهبعتلك النتايج فوراً:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("❌ إلغاء", callback_data="back:semesters")
            ]])
        )

    # Profile
    elif data == "profile":
        try:
            p = db_get_user_profile(uid)
            if not p:
                await q.edit_message_text("⚠️ مفيش بيانات.", reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔙 رجوع", callback_data="back:semesters")
                ]]))
                return
            u      = p["user"]
            name   = (u["name"] or "—").replace("*","").replace("`","").replace("_","")
            uname  = (f"@{u['username']}" if u["username"] else "—").replace("_", "\\_")
            joined = (u["joined_at"] or "")[:10]
            last   = (u["last_seen"] or "")[:16]
            avg    = p["avg_pct"]
            if avg == 0:       grade = "—"
            elif avg >= 85:    grade = "🏆 ممتاز"
            elif avg >= 70:    grade = "🥇 كويس جداً"
            elif avg >= 60:    grade = "👍 كويس"
            else:              grade = "📚 يحتاج مراجعة"
            best = ""
            if p["best_quiz"]:
                bq  = p["best_quiz"]
                pct = round(bq["score"] / bq["total"] * 100) if bq["total"] else 0
                btitle = bq["title"][:30].replace("*","").replace("`","").replace("_","")
                best = f"\n🏅 أحسن كويز: {btitle} — {pct}%"
            text = (
                f"👤 *بروفايلك*\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"📛 الاسم: {name}\n"
                f"🔗 يوزرنيم: {uname}\n"
                f"📅 انضم: {joined}\n"
                f"🕐 آخر ظهور: {last}\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"🧠 محاولات الكويز: {p['quiz_attempts']}\n"
                f"📊 متوسط الدرجات: {avg}%  {grade}\n"
                f"🔖 المحفوظات: {p['bookmarks_count']}"
                f"{best}"
            )
            await q.edit_message_text(
                text, parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔙 رجوع", callback_data="back:semesters")
                ]])
            )
        except Exception as e:
            logging.error(f"Profile error for uid {uid}: {e}", exc_info=True)
            await q.edit_message_text(
                f"⚠️ حصل خطأ: {e}",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔙 رجوع", callback_data="back:semesters")
                ]])
            )

    # Quiz start
    elif data.startswith("quiz:"):
        _, lid, qidx = data.split(":", 2)
        lid  = int(lid)
        qidx = int(qidx)
        quizzes = db_get_quizzes(lid)
        if qidx == 0:
            context.user_data[f"quiz_score_{lid}"] = 0
        if qidx >= len(quizzes):
            await q.edit_message_text(
                "🎉 *أنهيت الكويز!*\n\nعظيم، استمر في الدراسة 💪",
                parse_mode="Markdown"
            )
            return
        quiz = quizzes[qidx]
        score_so_far = context.user_data.get(f"quiz_score_{lid}", 0)
        await q.edit_message_text(
            f"🧠 *السؤال {qidx+1} من {len(quizzes)}*  •  ✅ {score_so_far} صح\n"
            f"━━━━━━━━━━━━━━━━━━━━\n\n"
            f"{md_escape(quiz['question'])}",
            parse_mode="Markdown", reply_markup=kb_quiz(quiz, qidx)
        )

    elif data.startswith("qa:"):
        _, lid, qidx, chosen = data.split(":", 3)
        lid     = int(lid)
        qidx    = int(qidx)
        quizzes = db_get_quizzes(lid)
        quiz    = quizzes[qidx]
        correct = quiz["answer"].upper() == chosen.upper()
        score_key = f"quiz_score_{lid}"
        if correct:
            context.user_data[score_key] = context.user_data.get(score_key, 0) + 1
        score  = context.user_data.get(score_key, 0)
        total  = len(quizzes)
        result = "✅ *إجابة صح!*" if correct else f"❌ *إجابة غلط!*\nالإجابة الصحيحة: *{md_escape(quiz['answer'])}*"
        next_q = qidx + 1
        if next_q < total:
            await q.edit_message_text(
                f"{result}\n\n_{md_escape(quiz['question'])}_",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("السؤال التالي ▶️", callback_data=f"quiz:{lid}:{next_q}")
                ]])
            )
        else:
            db_save_quiz_attempt(uid, lid, score, total)
            context.user_data.pop(score_key, None)
            pct = int(score / total * 100)
            if pct == 100:   grade = "🏆 ممتاز! إنت نجم!"
            elif pct >= 80:  grade = "🥇 كويس جداً!"
            elif pct >= 60:  grade = "👍 كويس، استمر!"
            elif pct >= 40:  grade = "📚 محتاج مراجعة أكتر"
            else:             grade = "💪 متستسلمش، ذاكر تاني!"
            attempts, avg = db_get_quiz_stats(lid)
            lec = db_get_lecture(lid)
            back_nid = nav_save(lec["semester"], lec["module"], "L", lec["subject"]) if lec else None
            kb_result = InlineKeyboardMarkup([
                [InlineKeyboardButton("🔄 حل تاني", callback_data=f"quiz:{lid}:0")],
                *([[InlineKeyboardButton("🔙 رجوع", callback_data=f"sub:{back_nid}:0")]] if back_nid else [])
            ])
            await q.edit_message_text(
                f"🎉 *خلصت الكويز!*\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"✅ إجباتك صح: *{score} من {total}*\n"
                f"📊 نسبتك: *{pct}%*\n"
                f"{result}\n\n"
                f"{grade}\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"👥 متوسط الطلاب: *{avg}%*  ({attempts} محاولة)",
                parse_mode="Markdown",
                reply_markup=kb_result
            )

    # Back buttons
    elif data.startswith("back:"):
        dest = data[5:]
        if dest == "semesters":
            await start(update, context)
        elif dest.startswith("modules:"):
            sem = dest[8:]
            await q.edit_message_text(
                f"📘 *{sem}*\nاختار الموديول 👇",
                parse_mode="Markdown", reply_markup=kb_modules(sem))
        elif dest.startswith("subjects:"):
            parts  = dest.split(":", 2)
            sem    = parts[1]
            module = parts[2]
            await q.edit_message_text(
                f"📘 {sem}  ›  📗 *{module}*\n\nاختار نوع المحتوى 👇",
                parse_mode="Markdown", reply_markup=kb_content_types(sem, module)
            )

    # ════════════════════════════════════════════════
    # ADMIN NAVIGATION  (Sem → Module → Subject → Lec)
    # ════════════════════════════════════════════════

    # ── Add Module / Add Subject (MUST be first — names contain _sem:/_mod:/_sub:) ──
    elif data.startswith("newmod_sem:"):
        sem = data.split(":", 1)[1]
        context.user_data.update({"step": "add_module_name", "new_module_sem": sem})
        await q.edit_message_text(
            f"📦 Adding module to *{sem}*\n\n✍️ Send the *module name*:",
            parse_mode="Markdown"
        )

    elif data.startswith("newsub_sem:"):
        sem = data.split(":", 1)[1]
        context.user_data["new_subject_sem"] = sem
        await q.edit_message_text(
            f"📂 *{sem}* — Choose a Module to add the subject in:",
            parse_mode="Markdown",
            reply_markup=kb_mods_for_new_subject(sem)
        )

    elif data == "newsub_back_sem":
        await q.edit_message_text(
            "📂 *Add Subject* — Choose Semester:",
            parse_mode="Markdown",
            reply_markup=kb_sems_for_new_subject()
        )

    elif data.startswith("newsub_mod:"):
        _, sem, module = data.split(":", 2)
        context.user_data.update({
            "step": "add_subject_name",
            "new_subject_sem": sem,
            "new_subject_module": module
        })
        await q.edit_message_text(
            f"📂 Adding subject to *{module}* ({sem})\n\n✍️ Send the *subject name*:",
            parse_mode="Markdown"
        )

    # Semester pickers
    elif data.endswith("_sem_back"):
        prefix = data.replace("_sem_back", "")
        await q.edit_message_text("📘 Choose Semester:", reply_markup=kb_admin_sems(prefix))

    elif "_sem:" in data:
        prefix, sem = data.split("_sem:", 1)
        await q.edit_message_text(
            f"📘 *{sem}* — Choose a Module:",
            parse_mode="Markdown", reply_markup=kb_admin_modules(prefix, sem)
        )

    # Module pickers
    elif data.endswith(tuple([f"_mod_back:{s}" for s in all_sems()])):
        idx    = data.rfind("_mod_back:")
        prefix = data[:idx]
        sem    = data[idx + len("_mod_back:"):]
        await q.edit_message_text(
            f"📘 *{sem}* — Choose a Module:",
            parse_mode="Markdown", reply_markup=kb_admin_modules(prefix, sem)
        )

    elif "_mod:" in data:
        prefix, rest = data.split("_mod:", 1)
        sem, module  = rest.split(":", 1)
        await q.edit_message_text(
            f"📘 {sem}  ›  📗 *{module}*\nChoose a Subject:",
            parse_mode="Markdown", reply_markup=kb_admin_subjects(prefix, sem, module)
        )

    # Subject pickers
    elif "_sub_back:" in data:
        idx    = data.rfind("_sub_back:")
        prefix = data[:idx]
        rest   = data[idx + len("_sub_back:"):]
        sem, module = rest.split(":", 1)
        await q.edit_message_text(
            f"📘 {sem}  ›  📗 *{module}*\nChoose a Subject:",
            parse_mode="Markdown", reply_markup=kb_admin_subjects(prefix, sem, module)
        )

    elif "_sub:" in data:
        prefix, rest          = data.split("_sub:", 1)
        sem, module, subject  = rest.split(":", 2)

        if prefix == "add":
            # content_type already stored from admin_add_ctype step
            ctype = context.user_data.get("content_type", "L")
            context.user_data.update({"sem": sem, "module": module, "subject": subject,
                                      "step": "add_title"})
            await q.edit_message_text(
                f"📘 {sem}  ›  📗 {module}  ›  📄 *{subject}*  ›  {ctype_label(ctype)}\n\n"
                f"✍️ *Step 4/4* — Send the *title*:",
                parse_mode="Markdown"
            )
        else:
            await q.edit_message_text(
                f"📘 {sem}  ›  📗 {module}  ›  📄 *{subject}*\nChoose a Lecture:",
                parse_mode="Markdown",
                reply_markup=kb_admin_lectures(prefix, sem, module, subject)
            )

    # Lecture pickers (edit / delete / quiz)
    elif "_lec:" in data:
        prefix, lid_str = data.split("_lec:", 1)
        lid = int(lid_str)
        lec = db_get_lecture(lid)

        if prefix == "del":
            db_delete_lecture(lid)
            await q.edit_message_text(f"✅ Deleted *{md_escape(lec['title'])}*.", parse_mode="Markdown")

        elif prefix == "edit":
            context.user_data["edit_lid"] = lid
            await q.edit_message_text(
                f"✏️ Editing: *{md_escape(lec['title'])}*\nWhat do you want to change?",
                parse_mode="Markdown", reply_markup=kb_edit_fields(lid)
            )

        elif prefix == "quiz":
            context.user_data["quiz_lid"] = lid
            context.user_data["step"]     = "quiz_question"
            await q.edit_message_text(
                f"🧠 Adding quiz to *{md_escape(lec['title'])}*\n\nSend the *question*:",
                parse_mode="Markdown"
            )

    # Admin: Add Content — type chosen, now pick semester
    elif data.startswith("admin_add_ctype:"):
        code = data.split(":", 1)[1]
        context.user_data["content_type"] = code
        await q.edit_message_text(
            f"➕ *Add {ctype_label(code)}*\n📘 *Step 2/4* — Choose Semester:",
            parse_mode="Markdown",
            reply_markup=kb_admin_sems("add")
        )

    # Edit field choice
    elif data.startswith("edf:"):
        _, lid, field = data.split(":", 2)
        context.user_data["edit_lid"]   = int(lid)
        context.user_data["edit_field"] = field
        context.user_data["step"]       = "awaiting_edit"
        prompts = {"title":"✍️ New title:","content":"📝 New content:","file":"📎 Send new file:"}
        await q.edit_message_text(prompts[field], parse_mode="Markdown")

    # ════════════════════════════════════════════════
    # EDIT / DELETE MODULE
    # ════════════════════════════════════════════════

    elif data.startswith("editmod_sem:"):
        sem = data.split(":", 1)[1]
        await q.edit_message_text(f"✏️ *Edit Module* — *{sem}*\nChoose module:",
                                  parse_mode="Markdown",
                                  reply_markup=kb_mods_for_edit_module(sem))

    elif data == "editmod_back_sem":
        await q.edit_message_text("✏️ *Edit Module* — Choose Semester:",
                                  parse_mode="Markdown",
                                  reply_markup=kb_sems_for_edit_module())

    elif data.startswith("editmod_pick:"):
        _, sem, module = data.split(":", 2)
        context.user_data.update({"step": "edit_module_name",
                                  "edit_mod_sem": sem, "edit_mod_old": module})
        await q.edit_message_text(
            f"✏️ Renaming module: *{module}*\n\nSend the *new name*:",
            parse_mode="Markdown"
        )

    elif data.startswith("delmod_sem:"):
        sem = data.split(":", 1)[1]
        await q.edit_message_text(f"🗑 *Delete Module* — *{sem}*\nChoose module:",
                                  parse_mode="Markdown",
                                  reply_markup=kb_mods_for_del_module(sem))

    elif data == "delmod_back_sem":
        await q.edit_message_text("🗑 *Delete Module* — Choose Semester:",
                                  parse_mode="Markdown",
                                  reply_markup=kb_sems_for_del_module())

    elif data.startswith("delmod_pick:"):
        _, sem, module = data.split(":", 2)
        cnt = db_module_count(sem, module)
        # Confirm button
        await q.edit_message_text(
            f"⚠️ Delete *{md_escape(module)}*?\nThis will remove *{cnt}* lecture(s) and all subjects inside.",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Confirm Delete", callback_data=f"delmod_confirm:{sem}:{module}")],
                [InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")]
            ])
        )

    elif data.startswith("delmod_confirm:"):
        _, sem, module = data.split(":", 2)
        db_delete_module(sem, module)
        await q.edit_message_text(f"✅ Module *{md_escape(module)}* deleted.", parse_mode="Markdown")

    # ════════════════════════════════════════════════
    # EDIT / DELETE SUBJECT
    # ════════════════════════════════════════════════

    elif data.startswith("editsub_sem:"):
        sem = data.split(":", 1)[1]
        await q.edit_message_text(f"✏️ *Edit Subject* — *{sem}*\nChoose module:",
                                  parse_mode="Markdown",
                                  reply_markup=kb_mods_for_edit_subject(sem))

    elif data == "editsub_back_sem":
        await q.edit_message_text("✏️ *Edit Subject* — Choose Semester:",
                                  parse_mode="Markdown",
                                  reply_markup=kb_sems_for_edit_subject())

    elif data.startswith("editsub_mod:"):
        _, sem, module = data.split(":", 2)
        context.user_data["editsub_sem"] = sem
        await q.edit_message_text(
            f"✏️ *{sem}* › *{module}* — Choose subject to rename:",
            parse_mode="Markdown", reply_markup=kb_subs_for_edit(sem, module)
        )

    elif data.startswith("editsub_back_mod:"):
        sem = data.split(":", 1)[1]
        await q.edit_message_text(f"✏️ *Edit Subject* — *{sem}*\nChoose module:",
                                  parse_mode="Markdown",
                                  reply_markup=kb_mods_for_edit_subject(sem))

    elif data.startswith("editsub_pick:"):
        _, sem, module, subject = data.split(":", 3)
        context.user_data.update({
            "step": "edit_subject_name",
            "edit_sub_sem": sem, "edit_sub_mod": module, "edit_sub_old": subject
        })
        await q.edit_message_text(
            f"✏️ Renaming subject: *{subject}*\n\nSend the *new name*:",
            parse_mode="Markdown"
        )

    elif data.startswith("delsub_sem:"):
        sem = data.split(":", 1)[1]
        await q.edit_message_text(f"🗑 *Delete Subject* — *{sem}*\nChoose module:",
                                  parse_mode="Markdown",
                                  reply_markup=kb_mods_for_del_subject(sem))

    elif data == "delsub_back_sem":
        await q.edit_message_text("🗑 *Delete Subject* — Choose Semester:",
                                  parse_mode="Markdown",
                                  reply_markup=kb_sems_for_del_subject())

    elif data.startswith("delsub_mod:"):
        _, sem, module = data.split(":", 2)
        await q.edit_message_text(
            f"🗑 *{sem}* › *{module}* — Choose subject to delete:",
            parse_mode="Markdown", reply_markup=kb_subs_for_del(sem, module)
        )

    elif data.startswith("delsub_back_mod:"):
        sem = data.split(":", 1)[1]
        await q.edit_message_text(f"🗑 *Delete Subject* — *{sem}*\nChoose module:",
                                  parse_mode="Markdown",
                                  reply_markup=kb_mods_for_del_subject(sem))

    elif data.startswith("delsub_pick:"):
        _, sem, module, subject = data.split(":", 3)
        cnt = db_subject_count(sem, module, subject)
        await q.edit_message_text(
            f"⚠️ Delete *{subject}*?\nThis will remove *{cnt}* lecture(s).",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Confirm Delete",
                                      callback_data=f"delsub_confirm:{sem}:{module}:{subject}")],
                [InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")]
            ])
        )

    elif data.startswith("delsub_confirm:"):
        _, sem, module, subject = data.split(":", 3)
        db_delete_subject(sem, module, subject)
        await q.edit_message_text(f"✅ Subject *{subject}* deleted.", parse_mode="Markdown")

    elif data == "admin:cancel":
        context.user_data.clear()
        await q.edit_message_text("❌ *تم الإلغاء.*", parse_mode="Markdown")

# ─────────────────────────────────────────────────────
# 💬  TEXT HANDLER
# ─────────────────────────────────────────────────────
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    uid  = update.effective_user.id
    ud   = context.user_data
    db_register(update.effective_user)

    if is_rate_limited(uid):
        await update.message.reply_text("⏳ Too fast! Please slow down.")
        return

    if text == "/cancel":
        context.user_data.clear()
        await update.message.reply_text("❌ Cancelled.", reply_markup=ReplyKeyboardRemove())
        return

    # Search from inline prompt
    if ud.get("step") == "searching":
        context.user_data.clear()
        await _do_search(update, context, text)
        return

    # Edit awaiting text
    if ud.get("step") == "awaiting_edit":
        field = ud.get("edit_field")
        lid   = ud.get("edit_lid")
        if field in ("title", "content"):
            db_update_lecture(lid, field, text)
            context.user_data.clear()
            await update.message.reply_text(
                f"✅ *{field.capitalize()}* updated!", parse_mode="Markdown", reply_markup=admin_kb()
            )
        return

    # Edit Module — rename step
    if ud.get("step") == "edit_module_name":
        sem        = ud.get("edit_mod_sem")
        old_module = ud.get("edit_mod_old")
        new_module = text.strip()
        if not new_module:
            await update.message.reply_text("⚠️ Please send a valid module name.")
            return
        db_rename_module(sem, old_module, new_module)
        context.user_data.clear()
        await update.message.reply_text(
            f"✅ Module renamed: *{md_escape(old_module)}* → *{md_escape(new_module)}*",
            parse_mode="Markdown", reply_markup=admin_kb()
        )
        return

    # Edit Subject — rename step
    if ud.get("step") == "edit_subject_name":
        sem         = ud.get("edit_sub_sem")
        module      = ud.get("edit_sub_mod")
        old_subject = ud.get("edit_sub_old")
        new_subject = text.strip()
        if not new_subject:
            await update.message.reply_text("⚠️ Please send a valid subject name.")
            return
        db_rename_subject(sem, module, old_subject, new_subject)
        context.user_data.clear()
        await update.message.reply_text(
            f"✅ Subject renamed: *{md_escape(old_subject)}* → *{md_escape(new_subject)}*",
            parse_mode="Markdown", reply_markup=admin_kb()
        )
        return

    # Add Module — name step
    if ud.get("step") == "add_module_name":
        sem = ud.get("new_module_sem")
        module_name = text.strip()
        if not module_name:
            await update.message.reply_text("⚠️ Please send a valid module name.")
            return
        db_add_module(sem, module_name)
        context.user_data.clear()
        await update.message.reply_text(
            f"✅ Module *{md_escape(module_name)}* added to *{md_escape(sem)}*! 🎉",
            parse_mode="Markdown", reply_markup=admin_kb()
        )
        return

    # Add Subject — name step
    if ud.get("step") == "add_subject_name":
        sem     = ud.get("new_subject_sem")
        module  = ud.get("new_subject_module")
        subject = text.strip()
        if not subject:
            await update.message.reply_text("⚠️ Please send a valid subject name.")
            return
        db_add_subject(sem, module, subject)
        context.user_data.clear()
        await update.message.reply_text(
            f"✅ Subject *{md_escape(subject)}* added to *{md_escape(module)}* ({md_escape(sem)})! 🎉",
            parse_mode="Markdown", reply_markup=admin_kb()
        )
        return

    # Add content — title step
    if ud.get("step") == "add_title":
        ud["title"] = text
        ud["step"]  = "add_content"
        ctype_code  = ud.get("content_type", "L")
        await update.message.reply_text(
            f"📝 *Step 4a* [{ctype_label(ctype_code)}] — Send the *content*:\n_(Markdown supported)_",
            parse_mode="Markdown"
        )
        return

    # Admin — add new admin
    if ud.get("step") == "add_admin_uid":
        try:
            new_aid = int(text.strip())
        except ValueError:
            await update.message.reply_text("⚠️ لازم يكون رقم. جرب تاني:")
            return
        if new_aid in ADMIN_IDS:
            context.user_data.clear()
            await update.message.reply_text(
                f"⚠️ الـ ID ده أدمن أصلي من الـ .env.",
                reply_markup=admin_kb()
            )
            return
        # جيب اسمه من users
        db   = get_db()
        user = db.execute("SELECT name FROM users WHERE user_id=?", (new_aid,)).fetchone()
        name = user["name"] if user else "Unknown"
        db_add_admin(new_aid, name)
        context.user_data.clear()
        await update.message.reply_text(
            f"✅ تم إضافة *{md_escape(name)}* (`{new_aid}`) كأدمن.",
            parse_mode="Markdown", reply_markup=admin_kb()
        )
        # إشعار الأدمن الجديد
        try:
            await context.bot.send_message(
                new_aid,
                "👑 *تم تعيينك كأدمن في البوت!*\n\nابعت /admin للدخول للوحة التحكم.",
                parse_mode="Markdown"
            )
        except: pass
        return

    # Admin — remove admin
    if ud.get("step") == "remove_admin_uid":
        try:
            rem_aid = int(text.strip())
        except ValueError:
            await update.message.reply_text("⚠️ لازم يكون رقم. جرب تاني:")
            return
        if rem_aid in ADMIN_IDS:
            context.user_data.clear()
            await update.message.reply_text(
                "❌ مش ممكن تحذف أدمن أصلي من الـ .env.",
                reply_markup=admin_kb()
            )
            return
        db   = get_db()
        user = db.execute("SELECT name FROM admins WHERE user_id=?", (rem_aid,)).fetchone()
        if not user:
            context.user_data.clear()
            await update.message.reply_text(
                f"⚠️ مفيش أدمن بالـ ID ده.", reply_markup=admin_kb()
            )
            return
        name = user["name"] or "—"
        db_remove_admin(rem_aid)
        context.user_data.clear()
        await update.message.reply_text(
            f"✅ تم حذف *{md_escape(name)}* (`{rem_aid}`) من الأدمنز.",
            parse_mode="Markdown", reply_markup=admin_kb()
        )
        try:
            await context.bot.send_message(rem_aid, "⚠️ تم إزالة صلاحيات الأدمن بتاعتك.")
        except: pass
        return

    # Admin — DM to single user: get ID
    if ud.get("step") == "dm_uid":
        try:
            target_uid = int(text.strip())
        except ValueError:
            await update.message.reply_text("⚠️ لازم يكون رقم. جرب تاني:")
            return
        ud["dm_target"] = target_uid
        ud["step"] = "dm_msg"
        await update.message.reply_text(
            f"✅ ID: `{target_uid}`\n\nدلوقتي ابعت الرسالة:",
            parse_mode="Markdown"
        )
        return

    # Admin — DM to single user: send message
    if ud.get("step") == "dm_msg":
        target_uid = ud.get("dm_target")
        context.user_data.clear()
        try:
            await context.bot.send_message(
                target_uid,
                f"📨 *رسالة من الإدارة*\n\n{text}",
                parse_mode="Markdown"
            )
            await update.message.reply_text(
                f"✅ اتبعتت للـ ID `{target_uid}` بنجاح.",
                parse_mode="Markdown", reply_markup=admin_kb()
            )
        except Exception as e:
            await update.message.reply_text(
                f"❌ فشل الإرسال: {e}", reply_markup=admin_kb()
            )
        return

    # Admin — search user by ID
    if ud.get("step") == "search_user":
        context.user_data.clear()
        query = text.strip()

        # لو رقم — بحث بالـ ID مباشرة
        if query.isdigit():
            search_uid = int(query)
            p = db_find_user(search_uid)
            if not p:
                await update.message.reply_text(
                    f"❌ مفيش يوزر بالـ ID: `{search_uid}`",
                    parse_mode="Markdown", reply_markup=admin_kb()
                )
                return
            await _send_user_result(update, context, p, search_uid)
            return

        # لو نص — fuzzy search على الأسماء واليوزرنيمز
        import difflib, unicodedata

        def normalize(s):
            """تطبيع النص: lowercase، شيل تشكيل، بدّل أحرف عربية شائعة."""
            if not s: return ""
            s = s.lower().strip()
            # إزالة التشكيل
            s = ''.join(c for c in unicodedata.normalize('NFD', s)
                        if unicodedata.category(c) != 'Mn')
            # توحيد بعض الأحرف العربية
            s = s.replace('أ','ا').replace('إ','ا').replace('آ','ا')
            s = s.replace('ة','ه').replace('ى','ي')
            return s

        # قاموس تقريبي للأسماء الشائعة (عربي/إنجليزي/دلع)
        NICKNAMES = {
            "medo": ["محمد", "mohamed", "muhammad", "mohammed"],
            "memo": ["محمد", "mohamed"],
            "hamada": ["حمادة", "حمد", "ahmed", "hamad"],
            "hamo":   ["حمادة", "حمد"],
            "soso":   ["سوسن", "سارة", "sara", "susan"],
            "nona":   ["نونا", "نورهان", "nour"],
            "toto":   ["توتو", "طارق", "tarek"],
            "bebo":   ["بيبو", "بسام", "bassam"],
            "koko":   ["كوكو", "كريم", "karim"],
            "lolo":   ["لولو", "لمياء", "لبنى"],
            "dodo":   ["دودو", "دينا", "دياب"],
            "roro":   ["رورو", "رنا", "رانيا", "rana"],
            "gogo":   ["جوجو", "جورج", "george"],
            "momo":   ["موموم", "محمد", "مصطفى"],
            "oso":    ["أسامة", "osama"],
            "salma":  ["سلمى", "salma"],
            "nour":   ["نور", "نورهان", "nour"],
        }

        db   = get_db()
        all_users = db.execute("SELECT user_id, name, username FROM users").fetchall()
        q_norm = normalize(query)

        scored = []
        for u in all_users:
            name_norm  = normalize(u["name"] or "")
            uname_norm = normalize(u["username"] or "")

            # نسبة التشابه المباشر
            r_name  = difflib.SequenceMatcher(None, q_norm, name_norm).ratio()
            r_uname = difflib.SequenceMatcher(None, q_norm, uname_norm).ratio()

            # contains check
            r_contains = 0.7 if (q_norm in name_norm or q_norm in uname_norm) else 0

            # nickname check
            nick_bonus = 0
            if q_norm in NICKNAMES:
                for variant in NICKNAMES[q_norm]:
                    if normalize(variant) in name_norm or normalize(variant) in uname_norm:
                        nick_bonus = 0.85
                        break

            score = max(r_name, r_uname, r_contains, nick_bonus)
            if score >= 0.45:
                scored.append((score, u))

        scored.sort(key=lambda x: x[0], reverse=True)
        results = scored[:5]

        if not results:
            await update.message.reply_text(
                f"❌ مفيش يوزر باسم *{query}*",
                parse_mode="Markdown", reply_markup=admin_kb()
            )
            return

        if len(results) == 1:
            # نتيجة واحدة — اعرضها مباشرة
            found_uid = results[0][1]["user_id"]
            p = db_find_user(found_uid)
            await _send_user_result(update, context, p, found_uid)
            return

        # أكتر من نتيجة — اعرض قايمة للاختيار
        lines = [f"🔎 نتايج البحث عن: {query}\n"]
        buttons = []
        for score, u in results:
            uname = f"@{u['username']}" if u["username"] else ""
            label = f"{u['name'] or '—'} {uname}".strip()
            lines.append(f"  • {label} — {u['user_id']}")
            buttons.append([InlineKeyboardButton(
                f"👤 {label[:30]}", callback_data=f"view_user:{u['user_id']}"
            )])
        await update.message.reply_text(
            "\n".join(lines),
            reply_markup=InlineKeyboardMarkup(buttons)
        )
        await update.message.reply_text("👆 اختار اليوزر", reply_markup=admin_kb())
        return

    # Add content — content step
    if ud.get("step") == "add_content":
        ud["content"] = text
        ud["step"]    = "add_file"
        await update.message.reply_text(
            "📎 *Step 4b* — Send a *file / photo / video* or tap Skip:",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup([["⏭ Skip"]], resize_keyboard=True)
        )
        return

    # Add lecture — skip file
    if ud.get("step") == "add_file" and text == "⏭ Skip":
        ctype_code = ud.get("content_type", "L")
        lid = db_add_lecture(ud["sem"], ud["module"], ud["subject"], ud["title"], ud["content"],
                             content_type=ctype_db(ctype_code))
        t, m, s = ud["title"], ud["module"], ud["subject"]
        label = ctype_label(ctype_code)
        context.user_data.clear()
        await update.message.reply_text(
            f"✅ *{t}* added to *{s}* ({m}) as {label}! ID: `{lid}`",
            parse_mode="Markdown", reply_markup=admin_kb()
        )
        return

    # Broadcast — message step
    if ud.get("step") == "broadcast_msg":
        ud["bc_msg"] = text
        ud["step"]   = "broadcast_time"
        await update.message.reply_text(
            "📅 Send *now* or schedule?\nType `now` or `YYYY-MM-DD HH:MM`:",
            parse_mode="Markdown"
        )
        return

    # Broadcast — time step
    if ud.get("step") == "broadcast_time":
        msg = ud.get("bc_msg", "")
        context.user_data.clear()
        if text.lower() == "now":
            sent = 0
            failed = 0
            for user_id in db_all_uids():
                try:
                    await context.bot.send_message(
                        chat_id=user_id,
                        text=f"📢 *Announcement*\n\n{msg}",
                        parse_mode="Markdown"
                    )
                    sent += 1
                except: 
                    failed += 1
                await asyncio.sleep(0.05)  # 20 msg/sec — تحت حد تيليجرام
            await update.message.reply_text(
                f"✅ Sent to *{sent}* users." + (f" ⚠️ Failed: {failed}" if failed else ""),
                parse_mode="Markdown", reply_markup=admin_kb()
            )
        else:
            try:
                datetime.strptime(text, "%Y-%m-%d %H:%M")
                db_schedule_broadcast(msg, text)
                await update.message.reply_text(
                    f"✅ Scheduled for `{text}`.", parse_mode="Markdown", reply_markup=admin_kb()
                )
            except ValueError:
                await update.message.reply_text("⚠️ Use `YYYY-MM-DD HH:MM` or `now`.",
                                                parse_mode="Markdown")
        return

    # Quiz — question
    if ud.get("step") == "quiz_question":
        ud["quiz_q"]    = text
        ud["quiz_opts"] = []
        ud["step"]      = "quiz_opts"
        await update.message.reply_text("✏️ Send *Option A*:", parse_mode="Markdown")
        return

    # Quiz — options
    if ud.get("step") == "quiz_opts":
        opts   = ud["quiz_opts"]
        labels = ["A","B","C","D"]
        if text.lower() == "skip" and len(opts) >= 2:
            ud["step"] = "quiz_answer"
            await update.message.reply_text(
                f"✅ Correct answer letter ({'/'.join(labels[:len(opts)])}):",
                parse_mode="Markdown"
            )
            return
        opts.append(text)
        if len(opts) < 4:
            await update.message.reply_text(
                f"✏️ Send *Option {labels[len(opts)]}* (or `skip` to stop):",
                parse_mode="Markdown"
            )
        else:
            ud["step"] = "quiz_answer"
            await update.message.reply_text("✅ Correct answer letter (A/B/C/D):",
                                            parse_mode="Markdown")
        return

    # Quiz — answer
    if ud.get("step") == "quiz_answer":
        answer = text.upper().strip()
        labels = ["A","B","C","D"]
        if answer not in labels[:len(ud.get("quiz_opts",[]))]:
            await update.message.reply_text("⚠️ Send a valid letter.")
            return
        db_add_quiz(ud["quiz_lid"], ud["quiz_q"], ud["quiz_opts"], answer)
        context.user_data.clear()
        await update.message.reply_text("✅ Quiz added!", reply_markup=admin_kb())
        return

    # Import Excel — module step
    if ud.get("step") == "import_module":
        if text not in all_modules():
            await update.message.reply_text("⚠️ Invalid module.")
            return
        ud["import_module"] = text
        ud["step"]          = "import_subject"
        sem = find_sem_for_module(text)
        rows = [[s] for s in db_get_subjects(sem, text)]
        await update.message.reply_text(
            f"📄 Choose the Subject in *{text}*:",
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardMarkup(rows, resize_keyboard=True)
        )
        return

    if ud.get("step") == "import_subject":
        if text not in all_subjects():
            await update.message.reply_text("⚠️ Invalid subject.")
            return
        ud["import_subject"] = text
        ud["step"]           = "import_file"
        await update.message.reply_text(
            f"📥 Now send the *.xlsx* file.\nRequired columns: `title` | `content` | `file_id` | `file_type`",
            parse_mode="Markdown", reply_markup=ReplyKeyboardRemove()
        )
        return

    # ── Admin panel buttons ───────────────────────
    if is_admin(uid):

        if text == "✏️ Edit Module":
            context.user_data.clear()
            await update.message.reply_text(
                "✏️ *Edit Module* — Choose Semester:",
                parse_mode="Markdown",
                reply_markup=kb_sems_for_edit_module()
            )
            return

        if text == "🗑 Delete Module":
            context.user_data.clear()
            await update.message.reply_text(
                "🗑 *Delete Module* — Choose Semester:",
                parse_mode="Markdown",
                reply_markup=kb_sems_for_del_module()
            )
            return

        if text == "✏️ Edit Subject":
            context.user_data.clear()
            await update.message.reply_text(
                "✏️ *Edit Subject* — Choose Semester:",
                parse_mode="Markdown",
                reply_markup=kb_sems_for_edit_subject()
            )
            return

        if text == "🗑 Delete Subject":
            context.user_data.clear()
            await update.message.reply_text(
                "🗑 *Delete Subject* — Choose Semester:",
                parse_mode="Markdown",
                reply_markup=kb_sems_for_del_subject()
            )
            return

        if text == "📦 Add Module":
            context.user_data.clear()
            await update.message.reply_text(
                "📦 *Add Module* — Choose Semester:",
                parse_mode="Markdown",
                reply_markup=kb_sems_for_new_module()
            )
            return

        if text == "📂 Add Subject":
            context.user_data.clear()
            await update.message.reply_text(
                "📂 *Add Subject* — Choose Semester:",
                parse_mode="Markdown",
                reply_markup=kb_sems_for_new_subject()
            )
            return

        if text == "➕ Add Content":
            context.user_data.clear()
            ctype_rows = [[InlineKeyboardButton(label, callback_data=f"admin_add_ctype:{code}")]
                          for label, code in CONTENT_TYPES]
            ctype_rows.append([InlineKeyboardButton("❌ Cancel", callback_data="admin:cancel")])
            await update.message.reply_text(
                "➕ *Add Content* — Choose type:",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup(ctype_rows)
            )
            return

        if text == "✏️ Edit Lecture":
            context.user_data.clear()
            await update.message.reply_text(
                "✏️ *Edit Lecture* — Choose Semester:",
                parse_mode="Markdown", reply_markup=kb_admin_sems("edit")
            )
            return

        if text == "🗑 Delete Lecture":
            context.user_data.clear()
            await update.message.reply_text(
                "🗑 *Delete Lecture* — Choose Semester:",
                parse_mode="Markdown", reply_markup=kb_admin_sems("del")
            )
            return

        if text == "🧠 Add Quiz":
            context.user_data.clear()
            await update.message.reply_text(
                "🧠 *Add Quiz* — Choose Semester:",
                parse_mode="Markdown", reply_markup=kb_admin_sems("quiz")
            )
            return

        if text == "💾 Backup":
            await update.message.reply_text("⏳ جاري عمل النسخة الاحتياطية…")
            try:
                stamp = datetime.now().strftime("%Y%m%d_%H%M")
                dst   = f"backup_{stamp}.db"
                shutil.copy2(DB_FILE, dst)
                await update.message.reply_text(
                    f"✅ تم الـ Backup: `{dst}`",
                    parse_mode="Markdown", reply_markup=admin_kb()
                )
            except Exception as e:
                await update.message.reply_text(f"❌ فشل الـ Backup: {e}", reply_markup=admin_kb())
            return

        if text == "📨 رسالة ليوزر":
            context.user_data.clear()
            context.user_data["step"] = "dm_uid"
            await update.message.reply_text(
                "📨 *رسالة ليوزر معين*\n\nابعت الـ Telegram ID:",
                parse_mode="Markdown", reply_markup=ReplyKeyboardRemove()
            )
            return

        if text == "📢 Broadcast":
            context.user_data.clear()
            context.user_data["step"] = "broadcast_msg"
            await update.message.reply_text(
                "📢 Send the broadcast *message*:", parse_mode="Markdown",
                reply_markup=ReplyKeyboardRemove()
            )
            return

        if text == "📊 Stats":
            u, l, top, by_mod, all_users = db_stats()
            # ── Summary message ──────────────────────────
            lines = ["📊 *Bot Statistics*\n",
                     f"👥 Total users: `{u}`",
                     f"📚 Total lectures: `{l}`\n",
                     "*Lectures per module:*"]
            for row in by_mod:
                lines.append(f"  └ {row['module']}: `{row['cnt']}`")
            if top:
                lines.append("\n*🔥 Top 5 viewed:*")
                for i, row in enumerate(top, 1):
                    lines.append(f"  {i}. {row['title']} ({row['subject']}) — `{row['views']}` views")
            await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
            # ── Users list (split into chunks of 30) ─────
            if all_users:
                chunk_size = 30
                for chunk_start in range(0, len(all_users), chunk_size):
                    chunk = all_users[chunk_start:chunk_start + chunk_size]
                    ulines = [f"👤 *Users {chunk_start+1}–{chunk_start+len(chunk)}:*\n"]
                    for usr in chunk:
                        uname = f"@{usr['username']}" if usr['username'] else "—"
                        joined = usr['joined_at'][:10] if usr['joined_at'] else "?"
                        ulines.append(
                            f"• `{usr['user_id']}` | {usr['name'] or '?'} | {uname} | 📅 {joined}"
                        )
                    await update.message.reply_text("\n".join(ulines), parse_mode="Markdown")
            return

        if text == "📥 Import Excel":
            context.user_data.clear()
            context.user_data["step"] = "import_module"
            rows = [[m] for m in all_modules()]
            await update.message.reply_text(
                "📥 *Import from Excel* — Choose Module:",
                parse_mode="Markdown",
                reply_markup=ReplyKeyboardMarkup(rows, resize_keyboard=True)
            )
            return

        if text == "📤 Export Excel":
            await export_cmd(update, context)
            return

        if text == "🏦 إدارة البنوك":
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton(f"📘 {s}", callback_data=f"adminbank_sem:{s}")]
                for s in all_sems()
            ] + [[InlineKeyboardButton("❌ إلغاء", callback_data="admin:cancel")]])
            await update.message.reply_text(
                "🏦 *إدارة بنوك الأسئلة*\nاختار السيميستر:",
                parse_mode="Markdown", reply_markup=kb
            )
            return

        if text == "🔤 ترتيب أبجدي":
            await update.message.reply_text("⏳ بيتم الترتيب الأبجدي لكل الملفات…")
            try:
                moved = db_sort_lectures_alpha()
                await update.message.reply_text(
                    f"✅ *تم الترتيب الأبجدي بنجاح!*\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"📂 عدد الملفات اللي اترتبت: *{moved}*\n"
                    f"🔤 كل المحاضرات دلوقتي مرتبة أبجدياً داخل كل مادة.",
                    parse_mode="Markdown",
                    reply_markup=admin_kb()
                )
            except Exception as e:
                await update.message.reply_text(f"❌ خطأ أثناء الترتيب: {e}", reply_markup=admin_kb())
            return

        if text == "📦 Import ZIP":
            context.user_data.clear()
            context.user_data["step"] = "import_zip"
            await update.message.reply_text(
                "📦 *Import ZIP*\n\n"
                "ابعت ملف ZIP بالـ structure ده:\n\n"
                "`Semester/Module/Subject/PREFIX_title.ext`\n\n"
                "البادئات:\n"
                "  `L_` — محاضرة\n"
                "  `P_` — عملي\n"
                "  `Q_` — بنك أسئلة\n"
                "  `S_` — ملخص\n\n"
                "مثال:\n"
                "`Semester1/Anatomy/Bones/L_intro.pdf`",
                parse_mode="Markdown",
                reply_markup=ReplyKeyboardMarkup([["❌ إلغاء"]], resize_keyboard=True)
            )
            return

        if text == "❌ إلغاء" and ud.get("step") == "import_zip":
            context.user_data.clear()
            await update.message.reply_text("❌ تم الإلغاء.", reply_markup=admin_kb())
            return

        if text == "👑 إضافة أدمن":
            context.user_data.clear()
            context.user_data["step"] = "add_admin_uid"
            # اعرض الأدمن الحاليين
            admins = db_list_admins()
            lines = ["👑 *الأدمنز الحاليين:*"]
            for a in admins:
                lines.append(f"  • {a['name'] or '—'} — `{a['user_id']}`")
            if not admins:
                lines.append("  لا يوجد أدمنز مضافين من البوت")
            lines.append("\nابعت الـ Telegram ID للأدمن الجديد:")
            await update.message.reply_text(
                "\n".join(lines), parse_mode="Markdown",
                reply_markup=ReplyKeyboardRemove()
            )
            return

        if text == "🗑 حذف أدمن":
            context.user_data.clear()
            admins = db_list_admins()
            if not admins:
                await update.message.reply_text(
                    "⚠️ مفيش أدمنز مضافين من البوت.", reply_markup=admin_kb()
                )
                return
            context.user_data["step"] = "remove_admin_uid"
            lines = ["🗑 *اختار أدمن تحذفه — ابعت الـ ID:*\n"]
            for a in admins:
                lines.append(f"  • {a['name'] or '—'} — `{a['user_id']}`")
            await update.message.reply_text(
                "\n".join(lines), parse_mode="Markdown",
                reply_markup=ReplyKeyboardRemove()
            )
            return

        if text == "🔎 بحث يوزر":
            context.user_data.clear()
            context.user_data["step"] = "search_user"
            await update.message.reply_text(
                "🔎 *بحث عن يوزر*\n\nابعت الـ Telegram ID:",
                parse_mode="Markdown",
                reply_markup=ReplyKeyboardRemove()
            )
            return

        if text == "🔙 Exit Admin":
            context.user_data.clear()
            await update.message.reply_text("👋 Exited.", reply_markup=ReplyKeyboardRemove())
            await start(update, context)
            return

    await update.message.reply_text("❓ Use /start to begin.")

# ─────────────────────────────────────────────────────
# 📎  FILE HANDLER
# ─────────────────────────────────────────────────────
async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ud  = context.user_data
    msg = update.message
    uid = update.effective_user.id

    def extract_file():
        if msg.photo:    return msg.photo[-1].file_id, "photo"
        if msg.video:    return msg.video.file_id, "video"
        if msg.document:
            fname = msg.document.file_name or ""
            if fname.lower().endswith(".pdf"):
                return msg.document.file_id, "pdf"
            return msg.document.file_id, "document"
        return None, None

    # Admin Bank Upload
    if ud.get("step") == "adminbank_upload":
        file_id, file_type = extract_file()
        if not file_id:
            await msg.reply_text("⚠️ ابعت ملف أو صورة أو فيديو.")
            return
        sem       = ud.get("bank_sem")
        exam_type = ud.get("bank_type")
        title     = msg.caption or (msg.document.file_name if msg.document else None) or f"ملف {exam_type}"
        db_add_exam_bank(sem, exam_type, title, file_id, file_type)
        label = "الميدتيرم" if exam_type == "midterm" else "الفاينال"
        ud.clear()
        await msg.reply_text(
            f"✅ *تم الرفع بنجاح!*\n📄 {title}\n🏦 بنك {label} — {sem}",
            parse_mode="Markdown", reply_markup=admin_kb()
        )
        return

    # ZIP import
    if ud.get("step") == "import_zip" and msg.document:
        fname = msg.document.file_name or ""
        if not fname.lower().endswith(".zip"):
            await msg.reply_text("⚠️ ابعت ملف .zip بس.")
            return

        # ── حد الحجم: 20 MB (حد تيليجرام Bot API) ──────────────
        MAX_ZIP_BYTES = 500 * 1024 * 1024  # 500 MB
        file_size = msg.document.file_size or 0
        if file_size > MAX_ZIP_BYTES:
            size_mb = file_size / (1024 * 1024)
            await msg.reply_text(
                f"❌ <b>الملف كبير أوي!</b>\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"📦 حجم الملف: <b>{size_mb:.1f} MB</b>\n"
                f"⚠️ الحد المسموح: <b>500 MB</b>\n\n"
                f"💡 <b>الحل:</b> قسّم الـ ZIP لأجزاء أصغر من 500 MB وابعتهم واحد واحد.",
                parse_mode="HTML",
                reply_markup=admin_kb()
            )
            context.user_data.clear()
            return

        CTYPE_MAP = {
            "l": "lectures", "p": "practicals",
            "q": "question_banks", "s": "summaries"
        }
        CTYPE_LABEL = {
            "lectures": "📖", "practicals": "🔬",
            "question_banks": "❓", "summaries": "📝"
        }

        await msg.reply_text("⏳ بيتحمل ويتحلل الـ ZIP…")
        try:
            import zipfile, tempfile, os, aiohttp
            tg_file = await msg.document.get_file()
            tmp_zip = f"import_{uid}.zip"

            # لو الـ Local API رجع مسار محلي، استخدمه مباشرة بدون تحميل
            if tg_file.file_path and tg_file.file_path.startswith("/"):
                local_zip_path = tg_file.file_path
                use_local = True
            else:
                # حمّل الملف بـ streaming عشان متملاش الـ RAM
                file_url = tg_file.file_path
                use_local = False
                async with aiohttp.ClientSession() as session:
                    async with session.get(file_url, timeout=aiohttp.ClientTimeout(total=600)) as resp:
                        with open(tmp_zip, "wb") as f:
                            async for chunk in resp.content.iter_chunked(1024 * 1024):
                                f.write(chunk)
                local_zip_path = tmp_zip

            added   = 0
            skipped = 0
            errors  = []
            db      = get_db()

            with zipfile.ZipFile(local_zip_path, "r") as zf:
                # فلتر الملفات بس وشيل __MACOSX
                members = [
                    m for m in zf.infolist()
                    if not m.is_dir()
                    and "__MACOSX" not in m.filename
                    and not os.path.basename(m.filename).startswith(".")
                ]
                # ترتيب أبجدي
                members.sort(key=lambda m: m.filename.lower())

                with tempfile.TemporaryDirectory() as tmpdir:
                    zf.extractall(tmpdir)

                    for member in members:
                        parts = member.filename.replace("\\", "/").split("/")
                        # نشيل أي فولدر root لو موجود
                        if len(parts) < 4:
                            skipped += 1
                            continue

                        # لو في root folder خارجي نشيله
                        if len(parts) == 5:
                            parts = parts[1:]

                        if len(parts) != 4:
                            skipped += 1
                            continue

                        sem, module, subject, filename = parts
                        sem     = sem.strip()
                        module  = module.strip()
                        subject = subject.strip()

                        # تحديد النوع من البادئة
                        base = os.path.splitext(filename)[0]
                        prefix = base[:2].lower() if len(base) >= 2 else ""
                        if prefix[0:1] in CTYPE_MAP and (len(base) < 2 or base[1:2] == "_"):
                            ctype  = CTYPE_MAP[prefix[0]]
                            title  = base[2:].strip() or base
                        else:
                            ctype = "lectures"
                            title = base.strip()

                        if not title:
                            skipped += 1
                            continue

                        # رفع الملف لتيليجرام عشان نجيب file_id
                        fpath = os.path.join(tmpdir, member.filename.replace("\\", "/"))
                        if not os.path.exists(fpath):
                            skipped += 1
                            continue

                        ext   = os.path.splitext(filename)[1].lower()
                        ftype = (
                            "pdf"      if ext == ".pdf"  else
                            "video"    if ext in (".mp4", ".mov", ".avi", ".mkv") else
                            "photo"    if ext in (".jpg", ".jpeg", ".png", ".webp") else
                            "document"
                        )

                        try:
                            with open(fpath, "rb") as f:
                                if ftype == "photo":
                                    sent = await context.bot.send_photo(uid, f)
                                    fid  = sent.photo[-1].file_id
                                elif ftype == "video":
                                    sent = await context.bot.send_video(uid, f)
                                    fid  = sent.video.file_id
                                else:
                                    sent = await context.bot.send_document(uid, f)
                                    fid  = sent.document.file_id
                        except Exception as e:
                            errors.append(f"{filename}: {e}")
                            skipped += 1
                            continue

                        # تسجيل الـ subject لو مش موجود
                        exists = db.execute(
                            "SELECT 1 FROM custom_subjects WHERE semester=? AND module=? AND subject=?",
                            (sem, module, subject)
                        ).fetchone()
                        if not exists:
                            # تأكد إن المحتوى مش في DATA الـ hardcoded
                            subs_in_data = DATA.get(sem, {}).get(module, [])
                            if subject not in subs_in_data:
                                db.execute(
                                    "INSERT OR IGNORE INTO custom_subjects (semester,module,subject) VALUES (?,?,?)",
                                    (sem, module, subject)
                                )

                        db_add_lecture(sem, module, subject, title, "", fid, ftype, ctype)
                        added += 1
                        await asyncio.sleep(0.05)

            db.commit()
            if not use_local and os.path.exists(tmp_zip):
                os.remove(tmp_zip)
            context.user_data.clear()

            summary = (
                f"✅ *اكتمل الـ Import!*\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"📥 اتضاف: *{added}* ملف\n"
                f"⏭ اتخطى: *{skipped}* ملف\n"
            )
            if errors:
                summary += f"⚠️ أخطاء: {len(errors)}\n"
                summary += "\n".join(f"  • {e}" for e in errors[:5])
            await msg.reply_text(summary, parse_mode="Markdown", reply_markup=admin_kb())

        except zipfile.BadZipFile:
            context.user_data.clear()
            await msg.reply_text("❌ الملف مش ZIP صحيح.", reply_markup=admin_kb())
        except Exception as e:
            context.user_data.clear()
            await msg.reply_text(f"❌ خطأ: {e}", reply_markup=admin_kb())
        return

    # Add lecture file
    if ud.get("step") == "add_file":
        fid, ftype = extract_file()
        if not fid:
            await msg.reply_text("⚠️ Send a photo, video, document, or PDF.")
            return
        ctype_code = ud.get("content_type", "L")
        lid = db_add_lecture(ud["sem"], ud["module"], ud["subject"],
                             ud["title"], ud["content"], fid, ftype, ctype_db(ctype_code))
        t, m, s = ud["title"], ud["module"], ud["subject"]
        label = ctype_label(ctype_code)
        context.user_data.clear()
        await msg.reply_text(
            f"✅ *{t}* added to *{s}* ({m}) as {label} with {ftype}! 🎉 ID: `{lid}`",
            parse_mode="Markdown", reply_markup=admin_kb()
        )
        return

    # Edit file
    if ud.get("step") == "awaiting_edit" and ud.get("edit_field") == "file":
        fid, ftype = extract_file()
        if not fid:
            await msg.reply_text("⚠️ Send a photo, video, document, or PDF.")
            return
        lid = ud["edit_lid"]
        db_update_lecture(lid, "file_id",   fid)
        db_update_lecture(lid, "file_type", ftype)
        context.user_data.clear()
        await msg.reply_text("✅ File updated!", reply_markup=admin_kb())
        return

    # Excel import — only handle xlsx here, ignore everything else
    if ud.get("step") == "import_file" and msg.document:
        fname = msg.document.file_name or ""
        if not fname.endswith(".xlsx"):
            await msg.reply_text("⚠️ Please send a .xlsx file.")
            return
        try:
            import openpyxl
            tg_file  = await msg.document.get_file()
            tmp      = f"import_{uid}.xlsx"
            await tg_file.download_to_drive(tmp)
            wb      = openpyxl.load_workbook(tmp)
            ws      = wb.active
            headers = [str(c.value or "").lower().strip()
                       for c in next(ws.iter_rows(min_row=1, max_row=1))]
            module  = ud["import_module"]
            subject = ud["import_subject"]
            sem     = find_sem_for_module(module)
            count   = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                r = dict(zip(headers, row))
                if not r.get("title") or not r.get("content"):
                    continue
                db_add_lecture(sem, module, subject, str(r["title"]), str(r["content"]),
                               str(r.get("file_id") or "") or None,
                               str(r.get("file_type") or "") or None)
                count += 1
            context.user_data.clear()
            await msg.reply_text(
                f"✅ Imported *{count}* lectures into *{subject}* ({module})!",
                parse_mode="Markdown", reply_markup=admin_kb()
            )
        except ImportError:
            await msg.reply_text("⚠️ Install openpyxl: `pip install openpyxl`",
                                 parse_mode="Markdown")
        except Exception as e:
            await msg.reply_text(f"❌ Import failed: {e}")

# ─────────────────────────────────────────────────────
# 📊  EXPORT REPORT  — /export
# ─────────────────────────────────────────────────────
def build_report_excel() -> bytes:
    """Build a full bot report as an Excel workbook and return raw bytes."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side, GradientFill)
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl غير مثبت — شغّل: pip install openpyxl")

    wb = Workbook()

    # ── shared styles ────────────────────────────────
    DARK       = "1A2A3A"
    GOLD       = "C9A84C"
    LIGHT_GOLD = "FFF3CD"
    WHITE      = "FFFFFF"
    GREEN_BG   = "D6F5E3"
    RED_BG     = "FDECEA"
    HEADER_FONT = Font(name="Arial", bold=True, color=WHITE, size=11)
    TITLE_FONT  = Font(name="Arial", bold=True, color=DARK,  size=14)
    BODY_FONT   = Font(name="Arial", size=10)
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    HEADER_FILL = PatternFill("solid", fgColor=DARK)
    GOLD_FILL   = PatternFill("solid", fgColor=GOLD)
    ALT_FILL    = PatternFill("solid", fgColor="F2F6FA")
    thin        = Side(style="thin", color="CCCCCC")
    BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_row(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font   = HEADER_FONT
            cell.fill   = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER

    def style_data_row(ws, row, cols, alt=False):
        fill = ALT_FILL if alt else PatternFill("solid", fgColor=WHITE)
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font      = BODY_FONT
            cell.fill      = fill
            cell.alignment = LEFT
            cell.border    = BORDER

    def add_title(ws, title, subtitle="", cols=9):
        end_col = get_column_letter(cols)
        ws.merge_cells(f"A1:{end_col}1")
        t = ws["A1"]
        t.value     = title
        t.font      = Font(name="Arial", bold=True, color=DARK, size=16)
        t.alignment = CENTER
        t.fill      = PatternFill("solid", fgColor=LIGHT_GOLD)
        ws.row_dimensions[1].height = 32
        if subtitle:
            ws.merge_cells(f"A2:{end_col}2")
            s = ws["A2"]
            s.value     = subtitle
            s.font      = Font(name="Arial", italic=True, color="666666", size=10)
            s.alignment = CENTER
            ws.row_dimensions[2].height = 18

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ════════════════════════════════════════════════
    # SHEET 1 — Summary Dashboard
    # ════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "📊 ملخص"

    add_title(ws1, "🩺  DU - Medicine 2024 — تقرير شامل", f"تاريخ التصدير: {now_str}", cols=8)

    with get_db() as db:
        total_users   = db.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        total_lec     = db.execute("SELECT COUNT(*) FROM lectures").fetchone()[0]
        total_quizzes = db.execute("SELECT COUNT(*) FROM quizzes").fetchone()[0]
        total_bm      = db.execute("SELECT COUNT(*) FROM bookmarks").fetchone()[0]
        total_views   = db.execute("SELECT COALESCE(SUM(views),0) FROM analytics").fetchone()[0]
        by_ctype      = db.execute(
            "SELECT content_type, COUNT(*) cnt FROM lectures GROUP BY content_type"
        ).fetchall()
        by_sem        = db.execute(
            "SELECT semester, COUNT(*) cnt FROM lectures GROUP BY semester ORDER BY semester"
        ).fetchall()

    # KPI boxes — row 4
    ws1.row_dimensions[4].height = 14
    kpis = [
        ("👥 المستخدمين",  total_users,   "A", "B"),
        ("📚 المحتوى",     total_lec,     "C", "D"),
        ("🧠 الأسئلة",     total_quizzes, "E", "F"),
        ("👁 المشاهدات",   total_views,   "G", "H"),
    ]
    for label, val, c1, c2 in kpis:
        ws1.merge_cells(f"{c1}5:{c2}5")
        ws1.merge_cells(f"{c1}6:{c2}6")
        lbl_cell = ws1[f"{c1}5"]
        val_cell = ws1[f"{c1}6"]
        lbl_cell.value     = label
        lbl_cell.font      = Font(name="Arial", bold=True, color="555555", size=10)
        lbl_cell.alignment = CENTER
        lbl_cell.fill      = PatternFill("solid", fgColor=LIGHT_GOLD)
        lbl_cell.border    = BORDER
        val_cell.value     = val
        val_cell.font      = Font(name="Arial", bold=True, color=DARK, size=18)
        val_cell.alignment = CENTER
        val_cell.fill      = PatternFill("solid", fgColor=WHITE)
        val_cell.border    = BORDER
    ws1.row_dimensions[5].height = 22
    ws1.row_dimensions[6].height = 32

    # Content by type table — row 9
    ws1["A8"] = "📦 المحتوى حسب النوع"
    ws1["A8"].font = Font(name="Arial", bold=True, size=12, color=DARK)
    ws1.row_dimensions[8].height = 20

    type_labels = {"lectures":"📖 محاضرات","practicals":"🔬 عملي",
                   "question_banks":"❓ بنك أسئلة","summaries":"📝 ملخصات"}
    ws1["A9"] = "النوع";  ws1["B9"] = "العدد";  ws1["C9"] = "النسبة"
    style_header_row(ws1, 9, 3)
    ct_map = {r["content_type"]: r["cnt"] for r in by_ctype}
    for i, (code, label) in enumerate(type_labels.items(), 10):
        cnt = ct_map.get(code, 0)
        ws1.cell(i, 1, label).border = BORDER
        ws1.cell(i, 2, cnt).border   = BORDER
        ws1.cell(i, 3, f"=IF(B{i}=0,\"-\",TEXT(B{i}/SUM(B10:B13),\"0.0%\"))").border = BORDER
        style_data_row(ws1, i, 3, alt=(i % 2 == 0))
        ws1.cell(i, 1).font = BODY_FONT
    ws1.cell(14, 1, "الإجمالي").font = Font(name="Arial", bold=True, color=DARK)
    ws1.cell(14, 2, f"=SUM(B10:B13)").font = Font(name="Arial", bold=True)
    ws1.cell(14, 3, "100%").font = Font(name="Arial", bold=True)
    for c in range(1, 4):
        ws1.cell(14, c).fill   = PatternFill("solid", fgColor=LIGHT_GOLD)
        ws1.cell(14, c).border = BORDER

    # Content by semester — row 9, col E
    ws1["E8"] = "📘 المحتوى حسب الترم"
    ws1["E8"].font = Font(name="Arial", bold=True, size=12, color=DARK)
    ws1["E9"] = "الترم";  ws1["F9"] = "العدد"
    style_header_row(ws1, 9, 0)
    ws1["E9"].font = HEADER_FONT; ws1["E9"].fill = HEADER_FILL; ws1["E9"].border = BORDER
    ws1["F9"].font = HEADER_FONT; ws1["F9"].fill = HEADER_FILL; ws1["F9"].border = BORDER
    for i, row in enumerate(by_sem, 10):
        ws1.cell(i, 5, row["semester"]).border = BORDER
        ws1.cell(i, 6, row["cnt"]).border      = BORDER
        style_data_row(ws1, i, 0)
        ws1.cell(i, 5).font = BODY_FONT; ws1.cell(i, 5).fill = ALT_FILL if i%2==0 else PatternFill("solid", fgColor=WHITE)
        ws1.cell(i, 6).font = BODY_FONT; ws1.cell(i, 6).fill = ALT_FILL if i%2==0 else PatternFill("solid", fgColor=WHITE)

    set_col_widths(ws1, [18, 10, 12, 4, 20, 10, 4, 4])

    # ════════════════════════════════════════════════
    # SHEET 2 — Users
    # ════════════════════════════════════════════════
    ws2 = wb.create_sheet("👥 المستخدمين")
    add_title(ws2, "👥  قائمة المستخدمين", f"تصدير: {now_str}", cols=6)

    headers = ["#", "User ID", "الاسم", "يوزرنيم", "تاريخ التسجيل", "آخر ظهور"]
    for c, h in enumerate(headers, 1):
        ws2.cell(4, c, h)
    style_header_row(ws2, 4, len(headers))

    with get_db() as db:
        users = db.execute(
            "SELECT user_id, name, username, joined_at, last_seen FROM users ORDER BY joined_at DESC"
        ).fetchall()

    for i, u in enumerate(users, 5):
        ws2.cell(i, 1, i - 4)
        ws2.cell(i, 2, u["user_id"])
        ws2.cell(i, 3, u["name"] or "—")
        ws2.cell(i, 4, f"@{u['username']}" if u["username"] else "—")
        ws2.cell(i, 5, (u["joined_at"] or "")[:16])
        ws2.cell(i, 6, (u["last_seen"] or "")[:16])
        style_data_row(ws2, i, len(headers), alt=(i % 2 == 0))

    # Total row
    total_row = len(users) + 5
    ws2.cell(total_row, 1, "الإجمالي").font = Font(name="Arial", bold=True)
    ws2.cell(total_row, 2, f"=COUNTA(B5:B{total_row-1})").font = Font(name="Arial", bold=True)
    for c in range(1, len(headers) + 1):
        ws2.cell(total_row, c).fill   = PatternFill("solid", fgColor=LIGHT_GOLD)
        ws2.cell(total_row, c).border = BORDER

    set_col_widths(ws2, [5, 14, 22, 20, 20, 20])

    # ════════════════════════════════════════════════
    # SHEET 2b — User Activity Details
    # ════════════════════════════════════════════════
    ws2b = wb.create_sheet("📈 نشاط المستخدمين")
    add_title(ws2b, "📈  نشاط المستخدمين التفصيلي", f"تصدير: {now_str}", cols=9)

    headers2b = ["#", "User ID", "الاسم", "يوزرنيم", "محاولات الكويز",
                 "متوسط الدرجات %", "المحاضرات المحفوظة", "تاريخ التسجيل", "آخر ظهور"]
    for c, h in enumerate(headers2b, 1):
        ws2b.cell(4, c, h)
    style_header_row(ws2b, 4, len(headers2b))

    with get_db() as db:
        users_detail = db.execute("""
            SELECT u.user_id, u.name, u.username, u.joined_at, u.last_seen,
                   COUNT(DISTINCT qa.id) AS quiz_attempts,
                   ROUND(AVG(CASE WHEN qa.total > 0 THEN CAST(qa.score AS FLOAT)/qa.total*100 ELSE NULL END), 1) AS avg_pct,
                   COUNT(DISTINCT bm.lecture_id) AS bookmarks_count
            FROM users u
            LEFT JOIN quiz_attempts qa ON qa.user_id = u.user_id
            LEFT JOIN bookmarks     bm ON bm.user_id = u.user_id
            GROUP BY u.user_id
            ORDER BY quiz_attempts DESC
        """).fetchall()

    for i, u in enumerate(users_detail, 5):
        avg = u["avg_pct"]
        ws2b.cell(i, 1, i - 4)
        ws2b.cell(i, 2, u["user_id"])
        ws2b.cell(i, 3, u["name"] or "—")
        ws2b.cell(i, 4, f"@{u['username']}" if u["username"] else "—")
        ws2b.cell(i, 5, u["quiz_attempts"])
        avg_cell = ws2b.cell(i, 6, f"{avg}%" if avg is not None else "—")
        ws2b.cell(i, 7, u["bookmarks_count"])
        ws2b.cell(i, 8, (u["joined_at"]  or "")[:16])
        ws2b.cell(i, 9, (u["last_seen"]  or "")[:16])
        style_data_row(ws2b, i, len(headers2b), alt=(i % 2 == 0))
        if avg is not None:
            clr = GREEN_BG if avg >= 60 else RED_BG
            ws2b.cell(i, 6).fill = PatternFill("solid", fgColor=clr)
            ws2b.cell(i, 6).font = Font(name="Arial", bold=True,
                                        color="1A6B3C" if avg >= 60 else "C0392B")

    set_col_widths(ws2b, [5, 14, 22, 20, 16, 18, 18, 20, 20])

    # ════════════════════════════════════════════════
    # SHEET 3 — Content (Lectures + all types)
    # ════════════════════════════════════════════════
    ws3 = wb.create_sheet("📚 المحتوى")
    add_title(ws3, "📚  المحتوى الكامل", f"تصدير: {now_str}", cols=9)

    headers3 = ["#", "ID", "النوع", "الترم", "الموديول", "المادة", "العنوان", "المشاهدات", "تاريخ الإضافة"]
    for c, h in enumerate(headers3, 1):
        ws3.cell(4, c, h)
    style_header_row(ws3, 4, len(headers3))

    ctype_ar = {"lectures":"📖 محاضرات","practicals":"🔬 عملي",
                "question_banks":"❓ بنك أسئلة","summaries":"📝 ملخصات"}

    with get_db() as db:
        lectures = db.execute(
            """SELECT l.id, l.content_type, l.semester, l.module, l.subject,
                      l.title, l.created_at, COALESCE(a.views,0) views
               FROM lectures l
               LEFT JOIN analytics a ON a.lecture_id = l.id
               ORDER BY l.semester, l.module, l.subject, l.id"""
        ).fetchall()

    for i, lec in enumerate(lectures, 5):
        views = lec["views"]
        ws3.cell(i, 1, i - 4)
        ws3.cell(i, 2, lec["id"])
        ws3.cell(i, 3, ctype_ar.get(lec["content_type"], lec["content_type"]))
        ws3.cell(i, 4, lec["semester"])
        ws3.cell(i, 5, lec["module"])
        ws3.cell(i, 6, lec["subject"])
        ws3.cell(i, 7, lec["title"])
        ws3.cell(i, 8, views)
        ws3.cell(i, 9, (lec["created_at"] or "")[:16])
        style_data_row(ws3, i, len(headers3), alt=(i % 2 == 0))
        # Color high-views rows
        if views >= 100:
            ws3.cell(i, 8).fill = PatternFill("solid", fgColor=GREEN_BG)
            ws3.cell(i, 8).font = Font(name="Arial", bold=True, color="1A6B3C")

    set_col_widths(ws3, [5, 7, 16, 10, 28, 26, 36, 12, 18])

    # ════════════════════════════════════════════════
    # SHEET 4 — Top Viewed
    # ════════════════════════════════════════════════
    ws4 = wb.create_sheet("🔥 الأكثر مشاهدة")
    add_title(ws4, "🔥  الأكثر مشاهدةً — Top 20", f"تصدير: {now_str}", cols=7)

    headers4 = ["#", "العنوان", "المادة", "الموديول", "الترم", "النوع", "المشاهدات"]
    for c, h in enumerate(headers4, 1):
        ws4.cell(4, c, h)
    style_header_row(ws4, 4, len(headers4))

    with get_db() as db:
        top = db.execute(
            """SELECT l.title, l.subject, l.module, l.semester, l.content_type, a.views
               FROM analytics a JOIN lectures l ON l.id=a.lecture_id
               ORDER BY a.views DESC LIMIT 20"""
        ).fetchall()

    medals = {1:"🥇", 2:"🥈", 3:"🥉"}
    for i, row in enumerate(top, 5):
        rank = i - 4
        ws4.cell(i, 1, medals.get(rank, rank))
        ws4.cell(i, 2, row["title"])
        ws4.cell(i, 3, row["subject"])
        ws4.cell(i, 4, row["module"])
        ws4.cell(i, 5, row["semester"])
        ws4.cell(i, 6, ctype_ar.get(row["content_type"], row["content_type"]))
        ws4.cell(i, 7, row["views"])
        style_data_row(ws4, i, len(headers4), alt=(i % 2 == 0))
        if rank <= 3:
            for c in range(1, len(headers4) + 1):
                ws4.cell(i, c).fill = PatternFill("solid", fgColor=LIGHT_GOLD)
        ws4.cell(i, 7).font = Font(name="Arial", bold=True, color="C0392B")

    set_col_widths(ws4, [6, 36, 26, 28, 10, 16, 12])

    # ════════════════════════════════════════════════
    # SHEET 5 — Quizzes
    # ════════════════════════════════════════════════
    ws5 = wb.create_sheet("🧠 الأسئلة")
    add_title(ws5, "🧠  بنك الأسئلة الكامل", f"تصدير: {now_str}", cols=9)

    headers5 = ["#", "السؤال", "A", "B", "C", "D", "الإجابة", "المادة", "العنوان"]
    for c, h in enumerate(headers5, 1):
        ws5.cell(4, c, h)
    style_header_row(ws5, 4, len(headers5))

    with get_db() as db:
        quizzes = db.execute(
            """SELECT q.question, q.opt_a, q.opt_b, q.opt_c, q.opt_d, q.answer,
                      l.subject, l.title
               FROM quizzes q JOIN lectures l ON l.id=q.lecture_id
               ORDER BY l.subject, l.title"""
        ).fetchall()

    for i, q in enumerate(quizzes, 5):
        ws5.cell(i, 1, i - 4)
        ws5.cell(i, 2, q["question"])
        ws5.cell(i, 3, q["opt_a"] or "")
        ws5.cell(i, 4, q["opt_b"] or "")
        ws5.cell(i, 5, q["opt_c"] or "")
        ws5.cell(i, 6, q["opt_d"] or "")
        ans_cell = ws5.cell(i, 7, q["answer"])
        ans_cell.fill = PatternFill("solid", fgColor=GREEN_BG)
        ans_cell.font = Font(name="Arial", bold=True, color="1A6B3C")
        ws5.cell(i, 8, q["subject"])
        ws5.cell(i, 9, q["title"])
        style_data_row(ws5, i, len(headers5), alt=(i % 2 == 0))
        ws5.cell(i, 7).fill = PatternFill("solid", fgColor=GREEN_BG)
        ws5.cell(i, 7).font = Font(name="Arial", bold=True, color="1A6B3C")

    set_col_widths(ws5, [5, 40, 18, 18, 18, 18, 10, 24, 30])

    # ════════════════════════════════════════════════
    # SHEET 6 — محاولات الكويز
    # ════════════════════════════════════════════════
    ws6 = wb.create_sheet("📝 محاولات الكويز")
    add_title(ws6, "📝  محاولات الكويز — كل المحاولات", f"تصدير: {now_str}", cols=10)

    headers6 = ["#", "User ID", "اسم المستخدم", "عنوان المحاضرة", "المادة",
                "الدرجة", "من أصل", "النسبة %", "التقييم", "التاريخ"]
    for c, h in enumerate(headers6, 1):
        ws6.cell(4, c, h)
    style_header_row(ws6, 4, len(headers6))

    with get_db() as db:
        attempts = db.execute(
            """SELECT qa.user_id, u.name, u.username,
                      l.title, l.subject,
                      qa.score, qa.total, qa.taken_at
               FROM quiz_attempts qa
               JOIN users    u ON u.user_id    = qa.user_id
               JOIN lectures l ON l.id         = qa.lecture_id
               ORDER BY qa.taken_at DESC"""
        ).fetchall()

    for i, a in enumerate(attempts, 5):
        pct   = round(a["score"] / a["total"] * 100) if a["total"] else 0
        if pct == 100:  grade = "🏆 ممتاز"
        elif pct >= 80: grade = "🥇 كويس جداً"
        elif pct >= 60: grade = "👍 كويس"
        elif pct >= 40: grade = "📚 يحتاج مراجعة"
        else:           grade = "💪 ضعيف"
        ws6.cell(i, 1,  i - 4)
        ws6.cell(i, 2,  a["user_id"])
        ws6.cell(i, 3,  a["name"] or "—")
        ws6.cell(i, 4,  a["title"])
        ws6.cell(i, 5,  a["subject"])
        ws6.cell(i, 6,  a["score"])
        ws6.cell(i, 7,  a["total"])
        ws6.cell(i, 8,  f"{pct}%")
        ws6.cell(i, 9,  grade)
        ws6.cell(i, 10, (a["taken_at"] or "")[:16])
        style_data_row(ws6, i, len(headers6), alt=(i % 2 == 0))
        clr = GREEN_BG if pct >= 60 else RED_BG
        for ci in [6, 7, 8]:
            ws6.cell(i, ci).fill = PatternFill("solid", fgColor=clr)

    set_col_widths(ws6, [5, 12, 22, 34, 24, 8, 8, 10, 14, 18])

    # freeze panes on all sheets
    for ws in [ws2, ws2b, ws3, ws4, ws5, ws6]:
        ws.freeze_panes = ws.cell(5, 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_user_report_excel(target_uid: int) -> bytes:
    """Excel report for a single user."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db   = get_db()
    user = db.execute("SELECT * FROM users WHERE user_id=?", (target_uid,)).fetchone()
    if not user:
        raise ValueError("يوزر مش موجود")

    quiz_rows = db.execute(
        """SELECT qa.score, qa.total, qa.taken_at, l.title, l.subject, l.semester, l.module
           FROM quiz_attempts qa JOIN lectures l ON l.id=qa.lecture_id
           WHERE qa.user_id=? ORDER BY qa.taken_at DESC""", (target_uid,)
    ).fetchall()

    bm_rows = db.execute(
        """SELECT l.title, l.subject, l.semester, l.module, l.content_type
           FROM bookmarks bm JOIN lectures l ON l.id=bm.lecture_id
           WHERE bm.user_id=? ORDER BY l.subject""", (target_uid,)
    ).fetchall()

    wb  = Workbook()
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    DARK      = "1A1A2E"
    GOLD      = "C9A84C"
    LIGHT_GOLD= "FFF3CD"
    GREEN_BG  = "D4EDDA"
    RED_BG    = "F8D7DA"
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill("solid", fgColor=DARK)
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BORDER      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    def style_header(ws, row, cols):
        for c in range(1, cols+1):
            ws.cell(row, c).font      = HEADER_FONT
            ws.cell(row, c).fill      = HEADER_FILL
            ws.cell(row, c).alignment = CENTER
            ws.cell(row, c).border    = BORDER
        ws.row_dimensions[row].height = 22

    def style_data(ws, row, cols, alt=False):
        for c in range(1, cols+1):
            ws.cell(row, c).font      = Font(name="Arial", size=10)
            ws.cell(row, c).alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row, c).border    = BORDER
            if alt:
                ws.cell(row, c).fill = PatternFill("solid", fgColor="F5F5F5")

    def add_title(ws, title, sub, cols):
        ec = get_column_letter(cols)
        ws.merge_cells(f"A1:{ec}1")
        t = ws["A1"]
        t.value = title; t.font = Font(name="Arial", bold=True, color=DARK, size=15)
        t.alignment = CENTER; t.fill = PatternFill("solid", fgColor=LIGHT_GOLD)
        ws.row_dimensions[1].height = 30
        ws.merge_cells(f"A2:{ec}2")
        s = ws["A2"]
        s.value = sub; s.font = Font(name="Arial", italic=True, color="666666", size=10)
        s.alignment = CENTER; ws.row_dimensions[2].height = 16

    def set_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 1: معلومات اليوزر
    ws1 = wb.active
    ws1.title = "👤 بيانات اليوزر"
    name   = user["name"]    or "—"
    uname  = f"@{user['username']}" if user["username"] else "—"
    joined = (user["joined_at"] or "")[:16]
    last   = (user["last_seen"]  or "")[:16]
    avg_pct = 0
    if quiz_rows:
        scores = [r["score"]/r["total"]*100 for r in quiz_rows if r["total"]]
        avg_pct = round(sum(scores)/len(scores), 1) if scores else 0

    add_title(ws1, f"👤 تقرير يوزر: {name}", f"تصدير: {now_str}", cols=2)
    info = [
        ("🆔 Telegram ID",    str(user["user_id"])),
        ("📛 الاسم",          name),
        ("🔗 يوزرنيم",        uname),
        ("📅 تاريخ التسجيل",  joined),
        ("🕐 آخر ظهور",       last),
        ("🧠 محاولات الكويز", str(len(quiz_rows))),
        ("📊 متوسط الدرجات",  f"{avg_pct}%"),
        ("🔖 المحفوظات",      str(len(bm_rows))),
    ]
    for i, (k, v) in enumerate(info, 4):
        ws1.cell(i, 1, k).font  = Font(name="Arial", bold=True, size=11)
        ws1.cell(i, 1).fill     = PatternFill("solid", fgColor=LIGHT_GOLD)
        ws1.cell(i, 1).border   = BORDER
        ws1.cell(i, 2, v).font  = Font(name="Arial", size=11)
        ws1.cell(i, 2).border   = BORDER
        ws1.row_dimensions[i].height = 20
    set_widths(ws1, [28, 35])

    # ── Sheet 2: الكويزات
    ws2 = wb.create_sheet("🧠 الكويزات")
    add_title(ws2, "🧠 كل محاولات الكويز", f"تصدير: {now_str}", cols=7)
    h2 = ["#", "المحاضرة", "المادة", "الترم", "الدرجة", "من أصل", "النسبة %"]
    for c, h in enumerate(h2, 1): ws2.cell(4, c, h)
    style_header(ws2, 4, len(h2))
    for i, r in enumerate(quiz_rows, 5):
        pct = round(r["score"]/r["total"]*100) if r["total"] else 0
        ws2.cell(i, 1, i-4); ws2.cell(i, 2, r["title"])
        ws2.cell(i, 3, r["subject"]); ws2.cell(i, 4, r["semester"])
        ws2.cell(i, 5, r["score"]);   ws2.cell(i, 6, r["total"])
        pct_cell = ws2.cell(i, 7, f"{pct}%")
        style_data(ws2, i, len(h2), alt=(i%2==0))
        pct_cell.fill = PatternFill("solid", fgColor=GREEN_BG if pct>=60 else RED_BG)
        pct_cell.font = Font(name="Arial", bold=True,
                             color="1A6B3C" if pct>=60 else "C0392B")
    set_widths(ws2, [5, 35, 22, 14, 10, 10, 12])

    # ── Sheet 3: المحفوظات
    ws3 = wb.create_sheet("🔖 المحفوظات")
    add_title(ws3, "🔖 المحاضرات المحفوظة", f"تصدير: {now_str}", cols=5)
    h3 = ["#", "المحاضرة", "المادة", "الترم", "النوع"]
    for c, h in enumerate(h3, 1): ws3.cell(4, c, h)
    style_header(ws3, 4, len(h3))
    for i, r in enumerate(bm_rows, 5):
        ws3.cell(i, 1, i-4); ws3.cell(i, 2, r["title"])
        ws3.cell(i, 3, r["subject"]); ws3.cell(i, 4, r["semester"])
        ws3.cell(i, 5, r["content_type"])
        style_data(ws3, i, len(h3), alt=(i%2==0))
    set_widths(ws3, [5, 38, 22, 14, 12])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_user_report_pdf(target_uid: int) -> bytes:
    """PDF report — premium dark-gold design with full user stats."""
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, KeepTogether
    )
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus.flowables import Flowable
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os as _os

    # ── Arabic font registration ──────────────────────────
    _ARABIC_FONT = "Helvetica"
    _ARABIC_FONT_BOLD = "Helvetica-Bold"
    _arabic_candidates = [
        ("/usr/share/fonts/truetype/fonts-hosny-amiri/Amiri-Regular.ttf",
         "/usr/share/fonts/truetype/fonts-hosny-amiri/Amiri-Bold.ttf",
         "Amiri", "Amiri-Bold"),
        ("/usr/share/fonts/truetype/arabeyes/ae_AlArabiya.ttf",
         "/usr/share/fonts/truetype/arabeyes/ae_AlArabiya.ttf",
         "AeAlArabiya", "AeAlArabiya"),
        ("/usr/share/fonts/truetype/noto/NotoNaskhArabic-Regular.ttf",
         "/usr/share/fonts/truetype/noto/NotoNaskhArabic-Bold.ttf",
         "NotoArabic", "NotoArabic-Bold"),
    ]
    for reg_path, bold_path, fname, fbold in _arabic_candidates:
        try:
            if _os.path.exists(reg_path):
                pdfmetrics.registerFont(TTFont(fname, reg_path))
                if _os.path.exists(bold_path) and bold_path != reg_path:
                    pdfmetrics.registerFont(TTFont(fbold, bold_path))
                else:
                    fbold = fname
                _ARABIC_FONT = fname
                _ARABIC_FONT_BOLD = fbold
                break
        except Exception:
            pass

    # ── Palette ───────────────────────────────────────────
    C_BG       = colors.HexColor("#0D0D1A")   # header background
    C_DARK     = colors.HexColor("#12122A")   # section headers
    C_MID      = colors.HexColor("#1E1E3A")   # table header rows
    C_GOLD     = colors.HexColor("#D4AF37")   # gold accent
    C_GOLD2    = colors.HexColor("#F5E27A")   # light gold text
    C_GREEN    = colors.HexColor("#1E4D2B")   # pass background
    C_GREEN_TXT= colors.HexColor("#A8F0C0")
    C_RED      = colors.HexColor("#5C1A1A")   # fail background
    C_RED_TXT  = colors.HexColor("#FFAAAA")
    C_ROW1     = colors.HexColor("#F8F6FF")   # table alt row
    C_ROW2     = colors.white
    C_LABEL    = colors.HexColor("#2A2A50")   # info label cells
    C_GREY     = colors.HexColor("#888888")

    # ── Fancy header banner flowable ──────────────────────
    class HeaderBanner(Flowable):
        def __init__(self, width, name, uid_str, generated):
            super().__init__()
            self.banner_w  = width
            self.banner_h  = 3.8 * cm
            self.name      = name
            self.uid_str   = uid_str
            self.generated = generated

        def draw(self):
            c = self.canv
            w, h = self.banner_w, self.banner_h

            # dark background
            c.setFillColor(C_BG)
            c.roundRect(0, 0, w, h, 8, fill=1, stroke=0)

            # gold top border stripe
            c.setFillColor(C_GOLD)
            c.rect(0, h - 4, w, 4, fill=1, stroke=0)

            # gold bottom border stripe
            c.rect(0, 0, w, 3, fill=1, stroke=0)

            # decorative left accent bar
            c.setFillColor(C_GOLD)
            c.rect(0, 0, 6, h, fill=1, stroke=0)

            # title
            c.setFillColor(C_GOLD2)
            c.setFont("Helvetica-Bold", 18)
            c.drawString(18, h - 26, "🎓  DU - Medicine 2024 — Personal Report")

            # subtitle
            c.setFillColor(C_GREY)
            c.setFont("Helvetica", 9)
            c.drawString(18, h - 42, f"DU - Medicine 2024  |  Personal Progress Report  |  {self.generated}")

            # gold divider inside banner
            c.setStrokeColor(C_GOLD)
            c.setLineWidth(0.5)
            c.line(18, h - 50, w - 10, h - 50)

            # user name
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 13)
            c.drawString(18, 22, self.name)

            # user id
            c.setFillColor(C_GOLD)
            c.setFont("Helvetica", 9)
            c.drawRightString(w - 10, 22, f"ID: {self.uid_str}")

        def wrap(self, availW, availH):
            return self.banner_w, self.banner_h

    # ── Fetch data ────────────────────────────────────────
    db   = get_db()
    user = db.execute("SELECT * FROM users WHERE user_id=?", (target_uid,)).fetchone()
    if not user:
        raise ValueError("User not found")

    quiz_rows = db.execute(
        """SELECT qa.score, qa.total, qa.taken_at, l.title, l.subject, l.semester
           FROM quiz_attempts qa JOIN lectures l ON l.id=qa.lecture_id
           WHERE qa.user_id=? ORDER BY qa.taken_at DESC""", (target_uid,)
    ).fetchall()

    bm_rows = db.execute(
        """SELECT l.title, l.subject, l.semester, l.content_type
           FROM bookmarks bm JOIN lectures l ON l.id=bm.lecture_id
           WHERE bm.user_id=? ORDER BY l.semester, l.subject""", (target_uid,)
    ).fetchall()

    # best & worst quiz
    best_quiz  = max(quiz_rows, key=lambda r: r["score"]/r["total"] if r["total"] else 0) if quiz_rows else None
    worst_quiz = min(quiz_rows, key=lambda r: r["score"]/r["total"] if r["total"] else 1) if quiz_rows else None

    # per-subject avg
    subj_scores: dict = defaultdict(list)
    for r in quiz_rows:
        if r["total"]:
            subj_scores[r["subject"]].append(r["score"]/r["total"]*100)
    subj_avg = {s: round(sum(v)/len(v), 1) for s, v in subj_scores.items()}

    name    = user["name"]    or "—"
    uname   = f"@{user['username']}" if user["username"] else "—"
    joined  = (user["joined_at"] or "")[:16]
    last    = (user["last_seen"]  or "")[:16]
    scores  = [r["score"]/r["total"]*100 for r in quiz_rows if r["total"]]
    avg_pct = round(sum(scores)/len(scores), 1) if scores else 0
    pass_ct = sum(1 for s in scores if s >= 60)
    fail_ct = len(scores) - pass_ct

    # ── Styles ────────────────────────────────────────────
    styles   = getSampleStyleSheet()

    def make_style(name_, **kw):
        kw.setdefault("fontName", _ARABIC_FONT)
        return ParagraphStyle(name_, parent=styles["Normal"], **kw)

    sec_style = make_style("SEC", fontSize=11, textColor=C_GOLD2,
                           fontName=_ARABIC_FONT_BOLD, spaceBefore=14, spaceAfter=5,
                           backColor=C_DARK, leftIndent=-2, borderPad=6)
    n_style   = make_style("N",   fontSize=9,  textColor=colors.HexColor("#DDDDDD"))
    small     = make_style("SM",  fontSize=8,  textColor=C_GREY)

    # ── Doc ───────────────────────────────────────────────
    C_PAGE_BG = colors.HexColor("#0A0A18")   # full page dark background

    def _draw_page_bg(canvas, doc_):
        """Paint the full page background dark on every page."""
        canvas.saveState()
        canvas.setFillColor(C_PAGE_BG)
        canvas.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)
        canvas.restoreState()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=1.8*cm, leftMargin=1.8*cm,
                            topMargin=1.5*cm,   bottomMargin=1.8*cm)
    W = A4[0] - 3.6*cm   # usable width
    story = []

    # ── HEADER BANNER ─────────────────────────────────────
    story.append(HeaderBanner(W, name, str(user["user_id"]),
                              datetime.now().strftime("%Y-%m-%d %H:%M")))
    story.append(Spacer(1, 10))

    # ── SECTION helper ────────────────────────────────────
    def section(title):
        story.append(Spacer(1, 4))
        tbl = Table([[Paragraph(f"  {title}", sec_style)]], colWidths=[W])
        tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), C_DARK),
            ("LEFTPADDING",   (0,0), (-1,-1), 0),
            ("RIGHTPADDING",  (0,0), (-1,-1), 0),
            ("TOPPADDING",    (0,0), (-1,-1), 0),
            ("BOTTOMPADDING", (0,0), (-1,-1), 0),
            ("LINEBELOW",     (0,0), (-1,-1), 1.5, C_GOLD),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 6))

    # ── TABLE header style helper ─────────────────────────
    def tbl_header_style(ncols):
        return [
            ("BACKGROUND",    (0,0), (-1,0),  C_MID),
            ("TEXTCOLOR",     (0,0), (-1,0),  C_GOLD),
            ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,0),  9),
            ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
            ("FONTSIZE",      (0,1), (-1,-1), 8),
            ("TEXTCOLOR",     (0,1), (-1,-1), colors.HexColor("#222222")),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [C_ROW2, C_ROW1]),
            ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
            ("LINEBELOW",     (0,0), (-1,0),  1,   C_GOLD),
            ("LEFTPADDING",   (0,0), (-1,-1), 6),
            ("TOPPADDING",    (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("ALIGN",         (0,0), (-1,-1), "LEFT"),
        ]

    # ════════════════════════════════════════════════════
    # 1️⃣  USER INFO
    # ════════════════════════════════════════════════════
    section("👤  Account Info")

    def info_row(label, val):
        return [
            Paragraph(f"<b>{label}</b>", make_style(f"IL{label}", fontSize=9,
                      textColor=C_GOLD2, backColor=C_LABEL)),
            Paragraph(str(val), make_style(f"IV{label}", fontSize=9,
                      textColor=colors.HexColor("#222222"))),
        ]

    info_data = [
        info_row("🆔  Telegram ID",    user["user_id"]),
        info_row("👤  Name",           name),
        info_row("📛  Username",        uname),
        info_row("📅  Joined",  joined),
        info_row("🕐  Last Seen",        last),
        info_row("🧠  Quiz Attempts",  len(quiz_rows)),
        info_row("📊  Avg Score",   f"{avg_pct}%"),
        info_row("✅  Pass / ❌  Fail", f"{pass_ct} / {fail_ct}"),
        info_row("🔖  Bookmarks",       len(bm_rows)),
    ]
    info_tbl = Table(info_data, colWidths=[5.5*cm, W - 5.5*cm])
    info_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (0,-1), C_LABEL),
        ("TEXTCOLOR",     (0,0), (0,-1), C_GOLD2),
        ("FONTNAME",      (0,0), (-1,-1),"Helvetica"),
        ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor("#DDDDDD")),
        ("ROWBACKGROUNDS",(1,0), (1,-1), [C_ROW2, C_ROW1]),
        ("LEFTPADDING",   (0,0), (-1,-1), 8),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LINEAFTER",     (0,0), (0,-1),  1, C_GOLD),
    ]))
    story.append(info_tbl)

    # ════════════════════════════════════════════════════
    # 2️⃣  PERFORMANCE SUMMARY (KPI cards as a mini table)
    # ════════════════════════════════════════════════════
    if quiz_rows:
        section("📈  Performance Summary")

        best_pct  = round(best_quiz["score"]/best_quiz["total"]*100)  if best_quiz  and best_quiz["total"]  else 0
        worst_pct = round(worst_quiz["score"]/worst_quiz["total"]*100) if worst_quiz and worst_quiz["total"] else 0

        def kpi_cell(label, val, color=C_DARK):
            return Paragraph(
                f'<font size="7" color="#D4AF37">{label}</font><br/>'
                f'<font size="14"><b>{val}</b></font>',
                make_style(f"KPI{label}", fontSize=14, textColor=colors.white,
                           alignment=1, leading=18)
            )

        kpi_data = [[
            kpi_cell("Avg Score",  f"{avg_pct}%"),
            kpi_cell("Pass",           f"{pass_ct}  ✅"),
            kpi_cell("Fail",           f"{fail_ct}  ❌"),
            kpi_cell("Best Score",      f"{best_pct}%"),
            kpi_cell("Worst Score",       f"{worst_pct}%"),
        ]]
        kw = W / 5
        kpi_tbl = Table(kpi_data, colWidths=[kw]*5, rowHeights=[1.4*cm])
        kpi_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), C_MID),
            ("GRID",          (0,0), (-1,-1), 1,   C_GOLD),
            ("ALIGN",         (0,0), (-1,-1), "CENTER"),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("LEFTPADDING",   (0,0), (-1,-1), 4),
            ("RIGHTPADDING",  (0,0), (-1,-1), 4),
        ]))
        story.append(kpi_tbl)

        # per-subject breakdown
        if subj_avg:
            story.append(Spacer(1, 8))
            subj_data = [["Subject", "Avg %", "Grade"]]
            for subj, avg in sorted(subj_avg.items(), key=lambda x: -x[1]):
                rating = "Excellent 🌟" if avg >= 85 else ("Very Good ✅" if avg >= 70 else ("Good 👍" if avg >= 60 else "Needs Review ⚠️"))
                subj_data.append([subj[:35], f"{avg}%", rating])
            subj_tbl = Table(subj_data, colWidths=[9*cm, 3*cm, W-12*cm])
            s = tbl_header_style(3)
            for i, (_, avg) in enumerate(sorted(subj_avg.items(), key=lambda x: -x[1]), 1):
                bg = colors.HexColor("#E8F5E9") if avg >= 60 else colors.HexColor("#FFEBEE")
                s.append(("BACKGROUND", (2,i), (2,i), bg))
            subj_tbl.setStyle(TableStyle(s))
            story.append(subj_tbl)

    # ════════════════════════════════════════════════════
    # 3️⃣  QUIZ ATTEMPTS
    # ════════════════════════════════════════════════════
    section("🧠  Quiz Attempts")
    if quiz_rows:
        q_data = [["#", "Lecture", "Subject", "Semester", "Score", "%", "Date"]]
        for i, r in enumerate(quiz_rows, 1):
            pct = round(r["score"]/r["total"]*100) if r["total"] else 0
            q_data.append([
                str(i),
                r["title"][:32],
                r["subject"][:20],
                r["semester"],
                f"{r['score']}/{r['total']}",
                f"{pct}%",
                (r["taken_at"] or "")[:10],
            ])
        q_tbl = Table(q_data, colWidths=[0.7*cm, 5.5*cm, 3.5*cm, 2*cm, 1.8*cm, 1.5*cm, 2.3*cm])
        s = tbl_header_style(7)
        for i, r in enumerate(quiz_rows, 1):
            pct = round(r["score"]/r["total"]*100) if r["total"] else 0
            if pct >= 60:
                s += [("BACKGROUND", (5,i), (5,i), C_GREEN),
                      ("TEXTCOLOR",  (5,i), (5,i), C_GREEN_TXT)]
            else:
                s += [("BACKGROUND", (5,i), (5,i), C_RED),
                      ("TEXTCOLOR",  (5,i), (5,i), C_RED_TXT)]
        q_tbl.setStyle(TableStyle(s))
        story.append(q_tbl)
    else:
        story.append(Paragraph("No quiz attempts yet.", n_style))

    # ════════════════════════════════════════════════════
    # 4️⃣  BOOKMARKS
    # ════════════════════════════════════════════════════
    story.append(Spacer(1, 8))
    section("🔖  Saved Lectures")
    if bm_rows:
        bm_data = [["#", "Lecture", "Subject", "Semester", "Type"]]
        for i, r in enumerate(bm_rows, 1):
            bm_data.append([str(i), r["title"][:35], r["subject"][:22],
                            r["semester"], r["content_type"] or "lecture"])
        bm_tbl = Table(bm_data, colWidths=[0.7*cm, 6*cm, 4*cm, 2*cm, W-12.7*cm])
        bm_tbl.setStyle(TableStyle(tbl_header_style(5)))
        story.append(bm_tbl)
    else:
        story.append(Paragraph("No saved lectures yet.", n_style))

    # ── Footer ────────────────────────────────────────────
    story.append(Spacer(1, 16))
    story.append(HRFlowable(width="100%", thickness=1, color=C_GOLD))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f'<font color="#888888" size="8">DU - Medicine 2024  •  Auto-generated Report  •  {datetime.now().strftime("%Y-%m-%d %H:%M")}</font>',
        make_style("FT", alignment=1)
    ))

    doc.build(story, onFirstPage=_draw_page_bg, onLaterPages=_draw_page_bg)
    buf.seek(0)
    return buf.getvalue()


async def export_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin only: /export — sends full Excel report."""
    uid = update.effective_user.id
    if not is_admin(uid):
        await update.message.reply_text("⛔️ Admin only.")
        return
    await update.message.reply_text("⏳ بيتجهز الـ Excel…")
    try:
        xlsx  = build_report_excel()
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        await context.bot.send_document(
            chat_id=uid,
            document=io.BytesIO(xlsx),
            filename=f"report_{stamp}.xlsx",
            caption="📊 التقرير الشامل"
        )
    except Exception as e:
        await update.message.reply_text(f"❌ حصل خطأ: {e}")

async def sheet_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """يبعت للطالب PDF بإحصائياته الشخصية."""
    uid  = update.effective_user.id
    db_register(update.effective_user)
    msg  = update.message

    await msg.reply_text("⏳ بيتجهز الـ PDF بتاعك…")
    try:
        pdf   = build_user_report_pdf(uid)
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        await context.bot.send_document(
            chat_id=uid,
            document=io.BytesIO(pdf),
            filename=f"my_report_{stamp}.pdf",
            caption="📋 تقريرك الشخصي — كويزاتك ومحفوظاتك"
        )
    except Exception as e:
        await msg.reply_text(f"⚠️ حصل خطأ: {e}")


# ─────────────────────────────────────────────────────
# 💾  BACKGROUND JOBS
# ─────────────────────────────────────────────────────
async def daily_backup(context: ContextTypes.DEFAULT_TYPE):
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    dst   = f"backup_{stamp}.db"
    shutil.copy2(DB_FILE, dst)
    logging.info(f"✅ Backup: {dst}")
    for aid in ADMIN_IDS:
        try:
            await context.bot.send_message(aid, f"💾 Backup saved: `{dst}`",
                                           parse_mode="Markdown")
        except: pass

async def daily_stats(context: ContextTypes.DEFAULT_TYPE):
    """إحصائيات يومية تتبعت للأدمن كل يوم الصبح."""
    db = get_db()
    today = datetime.now().strftime("%Y-%m-%d")
    new_users = db.execute(
        "SELECT COUNT(*) FROM users WHERE joined_at LIKE ?", (f"{today}%",)
    ).fetchone()[0]
    total_users = db.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    quiz_today  = db.execute(
        "SELECT COUNT(*) FROM quiz_attempts WHERE taken_at LIKE ?", (f"{today}%",)
    ).fetchone()[0]
    top_lec = db.execute(
        """SELECT l.title, l.subject, a.views
           FROM analytics a JOIN lectures l ON l.id=a.lecture_id
           ORDER BY a.views DESC LIMIT 3"""
    ).fetchall()
    top_lines = "\n".join(
        f"  {i+1}. {r['title'][:25]} — {r['views']} مشاهدة"
        for i, r in enumerate(top_lec)
    ) or "  لا يوجد"
    msg = (
        f"📅 *إحصائيات يوم {today}*\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"👥 يوزرز جدد اليوم: *{new_users}*\n"
        f"👤 إجمالي المستخدمين: *{total_users}*\n"
        f"🧠 محاولات كويز اليوم: *{quiz_today}*\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"🔥 أكثر المحاضرات مشاهدة:\n{top_lines}"
    )
    for aid in ADMIN_IDS:
        try:
            await context.bot.send_message(aid, msg, parse_mode="Markdown")
        except: pass

async def check_broadcasts(context: ContextTypes.DEFAULT_TYPE):
    for row in db_pending_broadcasts():
        sent = 0
        failed = 0
        for uid in db_all_uids():
            try:
                await context.bot.send_message(
                    uid, f"📢 *Announcement*\n\n{row['message']}", parse_mode="Markdown"
                )
                sent += 1
            except:
                failed += 1
            await asyncio.sleep(0.05)  # 20 msg/sec — تحت حد تيليجرام
        db_mark_sent(row["id"])
        logging.info(f"📢 Broadcast #{row['id']} → {sent} users, failed: {failed}")

# ─────────────────────────────────────────────────────
# ▶️  MAIN
# ─────────────────────────────────────────────────────
def main():
    logging.basicConfig(format="%(asctime)s [%(levelname)s] %(message)s", level=logging.INFO)
    init_db()

    LOCAL_API_URL = os.getenv("LOCAL_API_URL", "http://127.0.0.1:8081/bot")
    app = (
        ApplicationBuilder()
        .token(TOKEN)
        .base_url(LOCAL_API_URL)
        .base_file_url(LOCAL_API_URL.replace("/bot", "/file/bot"))
        .local_mode(True)
        .read_timeout(300)
        .write_timeout(300)
        .connect_timeout(30)
        .build()
    )
    app.add_handler(CommandHandler("start",  start))
    app.add_handler(CommandHandler("admin",  admin_cmd))
    app.add_handler(CommandHandler("search", search_cmd))
    app.add_handler(CommandHandler("export", export_cmd))
    app.add_handler(CommandHandler("sheet",  sheet_cmd))
    app.add_handler(CallbackQueryHandler(callback_handler))
    app.add_handler(MessageHandler(
        filters.Document.ALL | filters.PHOTO | filters.VIDEO, handle_file
    ))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    jq = app.job_queue
    jq.run_daily(daily_backup,  time=dtime(hour=3,  minute=0))
    jq.run_daily(daily_stats,   time=dtime(hour=8,  minute=0))
    jq.run_repeating(check_broadcasts, interval=60, first=10)

    print("🤖 DU - Medicine 2024 v3 is running...")
    app.run_polling()

if __name__ == "__main__":
    main()
